#!/depot/Python/Python-3.6.2/bin/python
import sys
sys.path.append('jiradc.api')
import jiradc.api
from shutil import copyfile
import re
from pathlib import Path
import time
import argparse as ap
import getpass
import os,shutil
import os.path
from os import path
import subprocess
import openpyxl as xl
from datetime import datetime, timedelta, date, tzinfo, timezone

#---------------------------------------------------------------
#  Message at beginning of execution ...
#---------------------------------------------------------------
def header(author_name):
    print (f'----------------------------------------------------------------------------------')
    print (f'-I- Report issues in JIRA; assign to ' + author_name)
    print (f"-I-   Type= Task, Occurred At= Development, L1= DesignWare Cores  L2= Methodology" )
    print (f'----------------------------------------------------------------------------------')
    return
#---------------------------------------------------------------
#  Message at end of execution ...
#---------------------------------------------------------------
def footer(author_name):
    print (f'----------------------------------------------------------------------------------')
    print (f'-I- Report issues in JIRA; assign to ' + author_name)
    print (f"-I-   Type= Task, Occurred At= Development, L1= DesignWare Cores  L2= Methodology" )
    print (f'----------------------------------------------------------------------------------')
    return
#---------------------------------------------------------------
#  Adjust the JIRA's ...  
#---------------------------------------------------------------
def bugParser(args, session):
    bug = None
    if (args.bug is None or args.bug == ''):
        print("Error: Bug STAR ID is a required input.")
        sys.exit(0)
    else:
        bug = session.get_issue(args.bug)
        if bug is None: 
            print(f'Cannot get issue by input Bug ID: {args.bug}. Please Double check.')
            sys.exit(0)
    return bug

#---------------------------------------------------------------
#  Adjust the JIRA's ...  
#---------------------------------------------------------------
def templateValidator(args, session, id):
    template = None
    if (id is None or id == ''):
        print("Error: Missing Jira template ID. Please double check.")
        sys.exit(0)
    else:
        template = session.get_issue(id)
        if template is None: 
            print(f'Cannot get issue by pre-defined template ID: {id}. Please Double check.')
            sys.exit(0)
    return template


#---------------------------------------------------------------
#  Create the 8D STORY ... setup the fields using the BUG's
#      L1-L4, priority, etc. 
#---------------------------------------------------------------
def storyCreator(args, session, jiraBUG, template, DueDateLatency):
    # Known good in description: urls, color tags, basic bullet points
    
    issue = jiradc.api.Story(session)
    #Categories
    issue.issue_type = 'story'
    issue.summary = '8D : ' + jiraBUG.summary
    issue.project = 'P' + jiraBUG.product_l1
    issue.product_l1 = jiraBUG.product_l1
    issue.product_l2 = jiraBUG.product_l2 
    issue.product_l3 = jiraBUG.product_l3
    issue.product_l4 = jiraBUG.product_l4
    issue.priority = jiraBUG.priority
    #format: 2020-10-30
    due_date = datetime.today() + timedelta(weeks=DueDateLatency)
    due_date = due_date.strftime('%Y-%m-%d')
    issue.due_date = due_date
    # issue.description = '\n'.join(desc)
    #debug
    issue.description = template.description

    if args.debug:
        print(issue)
        return 'P80001562-67244'
    else :
        (issue_key, errors) = issue.create()
        if errors:
            for error in errors: 
                print(error)
            sys.exit(0)
        else:
            if issue_key is None: 
                print(f'Failed to create new issue. Please Double check arguments or try again.')
                sys.exit(0)
            else:
                print (f'-I- Cloned TEMPLATE {template.key} to 8D Story {issue_key}')
                return issue_key

#---------------------------------------------------------------
#  Adjust the JIRA's ... links
#---------------------------------------------------------------
def linkIssue(jiraTEMPLATE, jiraBUG, args, session, jiraStoryKey, linkType, comment):
    jiraSTORY = session.get_issue(jiraStoryKey)

    if jiraSTORY is None: 
        print(f'Error: Failed to link issues, destination does not exist: {jiraStoryKey}. Please Double check.')
        sys.exit(0)

    #  Python API from IT does not allow a 'split from' relationship.
    #      So, using 'split to' relationship requires establishing link 
    #      in the BUG rather than in the STORY.
    #  Goal:  Add linkage in the BUG indicating it was split to the STORY
    errors = jiraBUG.add_link(linkType, jiraStoryKey, comment)
    
    if errors:
        for error in errors:
            print(error)
        sys.exit(0)
    else:
        print (f'-I- Adding link : BUG {jiraBUG.key} is {linkType} STORY {jiraStoryKey}')
        # return

    # Create a link to same issues that TEMPLATE JIRA had linked ...
    #     1/29/2021 : as of today, the links in the 8D Story TEMPLATE are 
    #         there to help 8D lead CLONE future tickets they need filed
    #         List includes :  QMS review ticket, MCT review ticket, PAs etc
    objs = jiraTEMPLATE.links()
    for linkage in objs:
        errors = jiraSTORY.add_link('relates to', linkage.key)

    jiraSTORY.add_comment(comment)
    return

#---------------------------------------------------------------
#  Adjust the JIRA's ...  watchers
#---------------------------------------------------------------
def watchIssue(jiraTEMPLATE, jiraBUG, session, jiraStoryKey, MCT_8D_lead_id, MCT_8D_lead_name, comment):
    jiraSTORY = session.get_issue(jiraStoryKey)

    user_list = MCT_8D_lead_name
    errors = jiraSTORY.add_watcher(MCT_8D_lead_id)

    # Inherit same WATCHERs list found in TEMPLATE ...
    objs = jiraTEMPLATE.watchers()
    for user in objs:
        user_list = user_list + ", " + user.name
        errors = jiraSTORY.add_watcher(user.id)
    print (f"-I- Adding watchers from TEMPLATE: ", user_list)

    # Inherit same WATCHERs list found in B-STAR ...
    user_list =  getpass.getuser()
    objs = jiraBUG.watchers()
    for user in objs:
        user_list = user_list + ", " + user.name
        errors = jiraSTORY.add_watcher(user.id)

    print (f"-I- Adding watchers from BUG: ", user_list)

    jiraSTORY.add_comment(comment)
    return

#---------------------------------------------------------------
#  Adjust the JIRA's ... attachments
#---------------------------------------------------------------
def attachmentsIssue(jiraTEMPLATE, jiraBUG, args, session, jiraStoryKey, linkType, comment):
    jiraSTORY = session.get_issue(jiraStoryKey)
    # Inherit same WATCHERs list found in TEMPLATE ...
    # Add attachments ... must be in same directory as script, or add relative path info
    #     These files are excellent guides for 8D leads and avoid having them chase
    #     down materials in power point slides etc.
    template_attachments = ( 
                             "B-STAR-into-8D.jpg",
                             "8D-Archive-Guide.jpg",
                             "Life-of-an-8D.jpg",
                             "QMS-2021-8D-Template.pptx",
                             "8D-Customer-Summary-Template.pptx",
                           )
    for file in template_attachments:
        print (f"-I- Adding attachment : ", file)
        jiraSTORY.add_attachment(file)

    jiraSTORY.add_comment(comment)
    return

#---------------------------------------------------------------
#  Adjust the JIRA's ... assignee
#---------------------------------------------------------------
def assignIssue(session, jiraStoryKey, name_8d_specialist):
    jiraID = session.get_issue(jiraStoryKey)
    #  By default, assign 8D STORY to the PEM running script.
    #      This ensures it's not left (1) unassigned or (2) assigned to
    #      me (juliano) (3) left assigned to same person as the TEMPLATE

    username = getpass.getuser()
    comment = "Re-assigning from '" + username + " to 8D Specialist '" + name_8d_specialist + "'"

    # Ex:  juliano
    jiraID.assignee = name_8d_specialist
    jiraID.update(comment)
    print (f"-I-", comment)

    return 

#---------------------------------------------------------------
#  check_group 
#---------------------------------------------------------------
def check_group(group):
    valid_groups = {
        "hbm": "HBM",
        "ddr": "DDR",
        "serdes--consumer": "Serdes--Consumer",
        "serdes--enterprise": "Serdes--Enterprise",
    }
    str_not_found = "-E- Invalid 'group'!"
    scrubbed_group_name = valid_groups.get(group.lower(), str_not_found)
    if scrubbed_group_name == str_not_found:
        exit_msg = "-F- Invalid 'group' provided: "+ group + "\n-I- Valid options:\t" + str(valid_groups.values())
        sys.exit( exit_msg)

    return scrubbed_group_name

#---------------------------------------------------------------
#  check_domain 
#---------------------------------------------------------------
def check_domain(domain):
    valid_domains = {
        "ams": "AMS",
        "digital": "Digital",
        "di": "DI",
        "layout": "Layout",
        "hardening": "Hardening",
    }
    str_not_found = "-E- Invalid 'domain'!"
    scrubbed_domain_name = valid_domains.get(domain.lower(), str_not_found)
    if scrubbed_domain_name == str_not_found:
        exit_msg = "-F- Invalid 'domain' provided: "+ domain + "\n-I- Valid options:\t" + str(valid_domains.values())
        sys.exit( exit_msg)

    return scrubbed_domain_name

#---------------------------------------------------------------
# This def is used to upload 8d slides to sharepoint, and create a web link.
#---------------------------------------------------------------
def upload(use_dev_jira_server, JIRA_URL, jiraStoryKey, my_url, my_filename, my_password):
    jiraSTORY = session.get_issue(jiraStoryKey)

    # Step 1 ... 
    #    Create URL based on JIRA SERVER (qa vs prod)
    #    replace 'browse' with the rest API keywords
    #    insert the new STORY's jira key/ID
    JIRA_URL_BASE = re.sub('browse','rest/api/latest/issue', JIRA_URL)
    STORY_URL_REST_API = JIRA_URL_BASE + jiraSTORY.key + '/remotelink'
    # Step 2 ...
    #    Create friendly filenames for (1) URL (2) unix
    #    for (1) ... replace spaces with %20 so MS-Sharepoint is happy
    #    for (2) ... need to escape the spaces chars => '\ '
    my_orig_filename = my_filename
    my_url_friendly_filename  = re.sub('\s+','%20',my_orig_filename)
    my_unix_friendly_filename = re.sub('\s+','\ ',my_orig_filename)

    # Step 3 ... 
    # Copy QMS template based on 8D file name convention 
    template = "QMS-2021-8D-Template.pptx"
    shutil.copy(template , my_filename)
    my_full_url = my_url + my_url_friendly_filename

		# Make sure you have user's password ready, so slides can be uploaded
    username = getpass.getuser()
    if (username and not my_password):
       my_password = getpass.getpass("Password for " + user + " user: ")

    # Step 4 ... upload slides to 8D Archive
    #   --> if user is running in the PRODUCTION JIRA server, then upload slides to 8D archive
    #   --> if not, don't upload to the archive cause it's just testing.
    #print (f'curl --ntlm --user \''+username+':'+my_password+'\' -k --upload-file '+my_filename+" "+my_url )
    # Modify 'my_filename' to make it unix friendly.
    my_curl_command = 'curl --ntlm --user \''+username+':'+my_password+'\' -k --upload-file '+my_unix_friendly_filename+' '+my_url 

    if not use_dev_jira_server:
       print (f'Uploading file to 8D Archive:\n\t' + my_url + ''+my_orig_filename+"\n" )
       os.system( my_curl_command )

    #----- Not working yet due to IT permission issue -----#
    # Step 5 ... Check-in slides uploaded to 8D Archive
    #print (f'curl --data "PostBack=true&CheckinAction=ActionCheckin&KeepCheckout=0&CheckinDescription=Copy of QMS slide template" https://sp-sg/sites/msip-design/_layouts/checkin.aspx?Source='+my_url+'/Forms/AllItems.aspx&FileName='+my_url_friendly_filename)
    #os.system('curl --data "PostBack=true&CheckinAction=ActionCheckin&KeepCheckout=0&CheckinDescription=Copy of QMS slide template" https://sp-sg/sites/msip-design/_layouts/checkin.aspx?Source='+my_url+'/Forms/AllItems.aspx&FileName='+my_url_friendly_filename)
    #------------------------------------------------------#

    # Step 6 ... add web link & comment to JIRA
    print (f'-I- Adding link to 8D slides:\n\t'+my_url+my_url_friendly_filename)
    errors = jiraSTORY.add_web_link(my_url, title=my_orig_filename)
    comment = "-I- 8D slides (1) uploaded to 8D archive (2) but NOT checked-in (3) and Jira Web Link created ... by script"
    print( comment )
    jiraSTORY.add_comment(comment)

    # Step 7 ...
    #     cleanup ... remove 8D slides from unix path
    os.remove( my_filename )
    return

#---------------------------------------------------------------
#  Message at beginning of execution ...
#     process_workbook('MSIP_Owners_Leads_orig.xlsx', str_domain, str_group, jiraBUG.product_l2)
#---------------------------------------------------------------
def process_workbook (filename, sheet_name, team_name, product_l2_name):

    wb = xl.load_workbook(filename)

    print( "Product L2 name is '"+ product_l2_name+ "'...searching for 8D specialist in sheet: '"+ sheet_name+ "'\n")
    sheet = wb[sheet_name]
    team_name = team_name.strip()
    product_l2_name = product_l2_name.strip()
    row_val = 1
    valid_row = 0
    for col in range(1, sheet.max_column - 1):
        cell2 = sheet.cell(1, col)   
        if 'Product L2' in cell2.value:
            for row in range(2, sheet.max_row - 1):
                row_val += 1
                cell1 = sheet.cell(row, col)
                multi_product_l2_name = re.search(rf"{product_l2_name}", cell1.value)
                if multi_product_l2_name:
                    #print(row_val)
                    valid_row = 1
                    break
                    
        if f'8D Analysis Specialist {team_name}' in cell2.value:
            if not valid_row:
                print("Bad product l2 name. Exiting...")
                exit()
            assignee_name = sheet.cell(row_val, col)
            
            if assignee_name.value:
                multi_assignee = re.search(".*[/|,].*", str(assignee_name.value))
            else:
                print(f"No eng: make sure that for {product_l2_name} assignee was filled")
                break
            if multi_assignee:
                multi_assignee = re.split("\s*[/|,]\s*", assignee_name.value)
                k = 1
                a = []
                print("There are multiple assignees, please press number of engineer whom need to be assigned:")
                for i in multi_assignee:
                    assignee_name.value = i
                    print(f"({k}) - eng{k}: {assignee_name.value}")
                    a.append(i)
                    k += 1
                choosen_assignee_num = input("< ")
                for j in range(1, k):
                    if choosen_assignee_num == str(j):
                        assignee_name.value = a[j-1]
                        return assignee_name.value
                        break                   
            elif assignee_name:
                return assignee_name.value

#---------------------------------------------------------------
#  Main 
#---------------------------------------------------------------
if __name__ == "__main__":
    #---------------------------------
    # Setup the Defaults....
    #    Due Date=6wks from today
    #    jira template ID for DEV/PROD
    script_home  = "/slowfs/us01sgarc00021/jiraUtils"
    jira_db_DEV  = 'https://jqa.internal.synopsys.com/browse/'
    jira_db_PROD = 'https://jira.internal.synopsys.com/browse/'
    fname_MSIP_XLS   = 'MSIP_Owners_Leads_orig.xlsx'
    URL__MSIP_XLS    = 'https://synopsys.sharepoint.com/sites/sg-methodology/latest/Shared%20Documents/Documentation/MSIP_Owners_Leads.xlsx'
    URL__8D_archive  = 'https://sp-sg/sites/msip-design/msip_common/Shared%20Documents/8D%20Archive/' 
    MCT_8D_lead_id   = 'juliano'
    MCT_8D_lead_name = 'Patrick Juliano'
    jira_template_id_DEV  = 'P80001562-89585'
    jira_template_id_PROD = 'P80001562-89585'
    SG_MCT_TAT_for_8Ds__D1_to_D6 = dict({'4-Urgent':6,'3-High':6,'2-Medium':10,'1-Low':10})
    SG_MCT_TAT_for_8Ds__D1_to_D8 = dict({'4-Urgent':11,'3-High':13,'2-Medium':18,'1-Low':21})
    DueDateLatency = 6 # default to 6wks

    #---------------------------------
    #  Cmd Line Parse & Help Messages
    header(MCT_8D_lead_name)
    help_msg = 'Example: autoGen8D.py --bug' + ' P80001562-71584 --group DDR --domain AMS'
    parser = ap.ArgumentParser(description = help_msg )
    parser.add_argument('--password', type = str, required=False, help = "SNPS password (default: password obtained from line2 in file '~/jiralogin.txt'")
    parser.add_argument('--bug',    type = str, required=True, help = 'Jira ID of bug STAR (e.g. P80001562-67244)')
    parser.add_argument('--group',  type = str, required=True, help = 'Options: DDR | HBM | SerDes--Consumer | SerDes--Enterprise')
    parser.add_argument('--domain', type = str, required=True, help = 'Options: AMS | Digital | DI | Layout | Hardening')
    parser.add_argument('--dev',   help='to run in JIRA development database', required=False, action='store_true')
    parser.add_argument("--debug", help="Script debugging mode", required=False, action='store_true')
    args = parser.parse_args()

    user_current_dir = os.getcwd()
    os.chdir( script_home )
    #---------------------------------
    # There's 2 JIRA servers, one for dev/experiments, the other for production
    #  prod - production
    #  qa - development only
    home = str(Path.home())
    home = home + '/jiralogin.txt'
    if args.password: 
        password = args.password
        print( "Password was obtained from cmd line\n" )
    elif path.exists(home):
        file = open( home ) 
        cnt = 0
        for line in file: 
            cnt+=1
            if cnt == 2:
               password = line.rstrip()
        print( "Password was obtained from file: 'jiralogin.txt'\n" )
    else:
        user =  getpass.getuser()
        password = getpass.getpass("Enter Password for " + user + " user: ")
    #print( "Password=" + password + "\n")

    if args.dev: 
        print (f"-I- Using JIRA's development server image.")
        session = jiradc.api.Session("qa")
        templates = [ jira_template_id_DEV ]
        JIRA_URL =  jira_db_DEV
    else:
        print (f"-W- Using JIRA's production database.")
        session = jiradc.api.Session("prod")
        templates = [ jira_template_id_PROD ]
        JIRA_URL = jira_db_PROD
    print (f'')

    #---------------------------------
    # Check with user that all is in order ...
    str_domain = check_domain(args.domain)
    str_group  = check_group(args.group)

    # Use the str_group as the key to look-up the name of the SHEET
    #    that should be used in the MSIP XLS to find the 8D specialist. 
    MSIP_sheet_name_map = {
        "HBM": "DDR",
        "DDR": "DDR",
        "Serdes--Consumer": "SERDES",
        "Serdes--Enterprise": "SERDES"
    }
    xls_sheet_name = MSIP_sheet_name_map.get(str_group, 'str_not_found')
    #print("msip_xls sheet name=", xls_sheet_name )
    #print("msip_xls file  name=", script_home )

    jiraTEMPLATE = templateValidator(args, session, templates[0])
    
    jiraBUG = bugParser(args, session)
    if jiraBUG and args.debug: print(jiraBUG.key)
    #-------------- Lilit's Magical Code -------------------
    name_8d_specialist = process_workbook(fname_MSIP_XLS, xls_sheet_name, str_domain, jiraBUG.product_l2_name)
    print("8D Specialist for this Product and Domain is =>'" + name_8d_specialist +"'")
#    sys.exit();
    #-------------- Lilit's Magical Code -------------------

    # Now, define the DueDate Latency based on the official TAT tables published in SG AMS Meth Doc
    DueDateLatency = SG_MCT_TAT_for_8Ds__D1_to_D6[jiraBUG.priority]
    # Now, create the 8D STORY JIRA ticket
    jiraSTORY = storyCreator(args, session, jiraBUG, jiraTEMPLATE, DueDateLatency)

    # Now, Add links 
    linkType = 'split to'
    comment  = 'Issue linked by script'
    print (f'')
    linkIssue(jiraTEMPLATE, jiraBUG, args, session, jiraSTORY, 'split to', 'Issue linked by script')

    print (f'')
    comment  = 'Watchers adjusted by script'
    watchIssue(jiraTEMPLATE, jiraBUG, session, jiraSTORY, MCT_8D_lead_id, MCT_8D_lead_name, comment)

    print (f'')
    comment  = 'Attachments added by script'
    attachmentsIssue(jiraTEMPLATE, jiraBUG, args, session, jiraSTORY, linkType, comment)

    print (f'')
    assignIssue(session, jiraSTORY, name_8d_specialist)

    #----------------------------------------------------------------
    # The 8D Archive doesn't have a 'Layout' folder because there's not
    #    enough of them to justify the complexity of adding another folder
    #----------------------------------------------------------------
    if (str_domain == 'Layout' ):
       str_domain = 'AMS'

    URL__8D_archive = URL__8D_archive + str_domain + '/' + str_group + '/WIP/'
    file = "8D__" + jiraBUG.key + "--" + jiraBUG.summary + ".pptx"
    # Microsoft file system will reject the following characters, so remove them from the file name
    # see JIRA for bug report :  P10020416-26242
    file_rename = file
    file_rename = file_rename.replace(':', '')
    file_rename = file_rename.replace('*', '')
    file_rename = file_rename.replace('?', '')
    file_rename = file_rename.replace('"', '')
    file_rename = file_rename.replace('<', '')
    file_rename = file_rename.replace('>', '')
    file_rename = file_rename.replace('|', '')
    file_rename = file_rename.replace('[', '')
    file_rename = file_rename.replace(']', '')
    file_rename = file_rename.replace('&', 'and')

    print (f'')
    
    upload(args.dev, JIRA_URL, jiraSTORY, URL__8D_archive, file_rename, password)
    print (f'')
    print (f'-I- *New* 8D Story URL :\t\t\t', JIRA_URL.replace(' ', '') + jiraSTORY)
#    if not file_rename == file:
       #print (f"-W- B-STAR summary has characters incomptaible w/ Microsoft Sharepoint!")
       #print (f'-W- Removing all of these characters from Summary used in filename ":*?<>|[]"')

    print (f'-I- 8D slides are named :\t\t\t' + file_rename)
    print (f"\n-W- Check-in slides for 8D lead to use : \t" + URL__8D_archive)
    print (f'')
    footer(MCT_8D_lead_name)
    os.chdir( user_current_dir )

