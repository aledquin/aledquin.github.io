#!/depot/Python/Python-3.8.0/bin/python
###############################################################################
#
# Name    : ibis_correlation_report.py
# Author  :	Raneem Khalil
# Date    : 2022-06-09 13:41:46
# Purpose : Compares the PU/PD (for TX IBIS models) or
# ODT (for RX IBIS models) impedances against "target/ideal" impedance.
#
###############################################################################

import argparse
import atexit
import os
import pathlib
import sys
from typing import Dict, List
from openpyxl import load_workbook

# ----------------------------------#
bindir = str(pathlib.Path(__file__).resolve().parent)
sys.path.append(bindir + '/../lib/python/Util')
sys.path.append(bindir + '/../lib/python')
# ----------------------------------#

# from CommonHeader import LOW, NULL_VAL, MEDIUM, HIGH
from Messaging import iprint, fatal_error, hprint
# from Messaging import dprint, eprint, vhprint
from Messaging import create_logger, footer, header
# from Misc import find_keys_missing_in_dict, run_system_cmd
# from P4Utils import da_p4_add_to_changelist, da_p4_create_instance
import CommonHeader
from Misc import utils__script_usage_statistics, get_release_version

__author__ = 'Raneem Khalil'
__version__ = get_release_version()


###############################################################################
def parse_args():
    # Always include -v, and -d
    parser = argparse.ArgumentParser(
        description="Script compares the PU/PD (for TX IBIS models)""or ODT (for RX IBIS models) impedances against"" 'target/ideal' impedance.")
    parser.add_argument('-v', metavar='<#>', type=int,default=0, help='verbosity')
    parser.add_argument('-d', metavar='<#>', type=int, default=0,help='debug')

    # -------------------------------------
    parser.add_argument('-correlationFile', type=str, required=True,help='Specify input correlation spreadsheet.(required)')
    # -------------------------------------
    args = parser.parse_args()
    return args


def main():
    args = parse_args()
    filename = os.getcwd() + '/' + os.path.basename(__file__) + '.log'
    # Create log file
    logger = create_logger(filename)  # noqa F841
    # Register exit function
    atexit.register(footer)
    # Initalise shared variables and run main
    CommonHeader.init(args, __author__, __version__)
    header()
    failed_models = run_correlation_report(args.correlationFile)
    if (len(failed_models) == 0):
        hprint('NO FAILED MODELS DETECTED')
    else:
        hprint('NUMBER OF FAILED MODELS: ' + str(len(failed_models)))
        print('\n')
        iprint('FAILED MODELS:' + str(failed_models))
    iprint(f"Log file: {filename}")
    utils__script_usage_statistics("ibis_correlation_report", "2022ww23")


def run_correlation_report(correlation_file: str) -> List[Dict]:
    """Creates the correlation report"""
    if '.xlsx' not in correlation_file:
        fatal_error('Input file is not of correct format (.xlsx)')
    try:
        wb = load_workbook(correlation_file)
    except FileNotFoundError:
        fatal_error(f'Correlation file not found at path: {correlation_file}')

    impedance_list = read_excel_sheet(wb)
    start_row,start_col = find_ibis_information(impedance_list)
    ibis_columns, PuResis_col, PdResis_col, OdtResis_col = find_ibis_columns(
        impedance_list,start_row,start_col)
    data_arr = extract_ibis_impedances_from_spreadsheet(
        impedance_list, ibis_columns, PuResis_col, PdResis_col, OdtResis_col)
    return compare_impedances(data_arr)


def read_excel_sheet(wb):
    names = wb.sheetnames
    if 'Output' in names:
        i = names.index('Output')
        impedance_list = wb[names[i]]
    else:
        fatal_error('Spreadsheet of interest does not exist')

    return impedance_list


def find_ibis_information(impedance_list):
    start_row = 0
    start_col = 0
    for row in impedance_list.iter_rows():
        for cell in row:
            if str(cell.value) == 'IBIS':
                start_row = cell.row
                start_col = cell.column
    if ((start_col == 0) or (start_row == 0)):
        fatal_error('IBIS impedance information does not exist')

    return start_row, start_col


def find_ibis_columns(impedance_list,start_row,start_col):
    ibis_columns_names = []
    ibis_columns_titles = impedance_list.iter_rows(min_col=start_col, max_col=start_col + 3, min_row=start_row + 1, max_row=start_row + 1)
    for r in ibis_columns_titles:
        for c in r:
            ibis_columns_names.append(c.value)
            if str(c.value) == 'PU resis':
                PuResis_col = c.column
            if str(c.value) == 'PD resis':
                PdResis_col = c.column
            if str(c.value) == 'ODT resis':
                OdtResis_col = c.column
    if (('Case' not in ibis_columns_names) or (
        'PU resis' not in ibis_columns_names) or (
            'PD resis' not in ibis_columns_names) or (
                'ODT resis' not in ibis_columns_names)):
        fatal_error('One or more of the required columns is missing')
    ibis_columns = impedance_list.iter_rows(
        min_col=start_col, max_col=start_col + 3)

    return ibis_columns, PuResis_col, PdResis_col, OdtResis_col


def extract_ibis_impedances_from_spreadsheet(impedance_list, ibis_columns, PuResis_col, PdResis_col, OdtResis_col):
    data_arr = []
    for row in ibis_columns:
        for cell in row:
            if 'drv' in str(cell.value):
                data_dict = {}
                r = cell.row
                c = cell.column
                data_dict['Case'] = str(cell.value)
                data_dict['Ideal Impedance'] = float(impedance_list.cell(row=r + 1, column=c).value)
                data_dict['PU Impedance'] = impedance_list.cell(row=r + 1, column=PuResis_col).value
                data_dict['PD Impedance'] = impedance_list.cell(row=r + 1, column=PdResis_col).value
                data_arr.append(data_dict)
            elif 'rcv' in str(cell.value):
                data_dict = {}
                r = cell.row
                c = cell.column
                data_dict['Case'] = str(cell.value)
                if (((impedance_list.cell(row=r + 1, column=c).value)).isnumeric()):
                    data_dict['Ideal Impedance'] = float(impedance_list.cell(row=r + 1, column=c).value)
                else:
                    data_dict['Ideal Impedance'] = impedance_list.cell(row=r + 1, column=c).value
                data_dict['ODT Impedance'] = impedance_list.cell(row=r + 1, column=OdtResis_col).value
                data_arr.append(data_dict)

    return data_arr


def compare_impedances(data_arr):
    index = 0
    length = len(data_arr)
    failed_models = []

    while index < length:
        curr_dict = data_arr[index]
        if ('ODT Impedance' not in curr_dict.keys()):
            ideal_impedance = float(curr_dict['Ideal Impedance'])
            pu_impedance = float(curr_dict['PU Impedance'])
            pd_impedance = float(curr_dict['PD Impedance'])
            allowed_min = ideal_impedance - 0.1 * ideal_impedance
            allowed_max = ideal_impedance + 0.1 * ideal_impedance
            if ((pu_impedance <= allowed_max) and (pu_impedance >= allowed_min)):
                pass
            else:
                curr_failed_model = {}
                curr_failed_model['Case'] = curr_dict['Case']
                curr_failed_model['Ideal Impedance'] = ideal_impedance
                curr_failed_model['PU Impedance'] = pu_impedance
                failed_models.append(curr_failed_model)
                iprint('CASE' + '         ' + 'EXPECTED'' PU IMPEDANCE' + '         ' + 'MEASURED'' PU IMPEDANCE' + "\n" + str(curr_failed_model['Case']) + '                ' + str(curr_failed_model['Ideal Impedance']) + '                             ' + str(curr_failed_model['PU Impedance']))
                hprint('FAIL')
                print('\n')
            if ((pd_impedance <= allowed_max) and (pd_impedance >= allowed_min)):
                pass
            else:
                curr_failed_model = {}
                curr_failed_model['Case'] = curr_dict['Case']
                curr_failed_model['Ideal Impedance'] = ideal_impedance
                curr_failed_model['PD Impedance'] = pd_impedance
                failed_models.append(curr_failed_model)
                iprint('CASE' + '         ' + 'EXPECTED'' PD IMPEDANCE' + '         ' + 'MEASURED'' PD IMPEDANCE' + "\n" + str(curr_failed_model['Case']) + '                ' + str(curr_failed_model['Ideal Impedance']) + '                             ' + str(curr_failed_model['PD Impedance']))
                hprint('FAIL')
                print('\n')
        elif ('ODT Impedance' in curr_dict.keys()):
            ideal_impedance = curr_dict['Ideal Impedance']
            odt_impedance = float(curr_dict['ODT Impedance'])
            if (ideal_impedance == 'off'):
                if (odt_impedance >= 10000.0):
                    pass
                else:
                    curr_failed_model = {}
                    curr_failed_model['Case'] = curr_dict['Case']
                    curr_failed_model['Ideal Impedance'] = ideal_impedance
                    curr_failed_model['ODT Impedance'] = odt_impedance
                    failed_models.append(curr_failed_model)
                    print('\n')
                    iprint('CASE' + '         ' + 'EXPECTED'' ODT IMPEDANCE' + '         ' + 'MEASURED'' ODT IMPEDANCE' + "\n" + str(curr_failed_model['Case']) + '                ' + str(curr_failed_model['Ideal Impedance']) + '                             ' + str(curr_failed_model['ODT Impedance']))
                    hprint('FAIL')
            else:
                allowed_min = ideal_impedance - 0.1 * ideal_impedance
                allowed_max = ideal_impedance + 0.1 * ideal_impedance
                if ((odt_impedance <= allowed_max) and (odt_impedance >= allowed_min)):
                    pass
                else:
                    curr_failed_model = {}
                    curr_failed_model['Case'] = curr_dict['Case']
                    curr_failed_model['Ideal Impedance'] = ideal_impedance
                    curr_failed_model['ODT Impedance'] = odt_impedance
                    failed_models.append(curr_failed_model)
                    iprint('CASE' + '         ' + 'EXPECTED'' ODT IMPEDANCE' + '         ' + 'MEASURED'' ODT IMPEDANCE' + "\n" + str(curr_failed_model['Case']) + '                ' + str(curr_failed_model['Ideal Impedance']) + '                             ' + str(curr_failed_model['ODT Impedance']))
                    hprint('FAIL')
                    print('\n')
        index += 1

    return failed_models


if __name__ == '__main__':
    main()
