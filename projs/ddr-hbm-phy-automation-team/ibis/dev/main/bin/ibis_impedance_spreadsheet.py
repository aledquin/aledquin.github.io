#!/depot/Python/Python-3.8.0/bin/python
###############################################################################
# Name    : ibis_impedance_spreadsheet.py
# Author  : Forest Li
#         : Harsimrat Singh Wadhawan
# Date    : 2022-04-12 12:50:04
# Purpose : Check for IBIS impedance spreadhsheet and compare model names within it.
###############################################################################

import argparse
import atexit
import json
import os
import pathlib
from pickle import FALSE, TRUE
import re
import sys
import time
from typing import List, Dict

import inquirer
from openpyxl import load_workbook

# ---------------------------------- #
# ----------CUSTOM IMPORTS---------- #
# ---------------------------------- #
bindir = str(pathlib.Path(__file__).resolve().parent)
sys.path.append(bindir + "/../lib/python/Util")
sys.path.append(bindir + "/../lib/python")
# ---------------------------------- #

from CommonHeader import LOW, NULL_VAL, INSANE, HIGH, SUPER
from Messaging import fatal_error, dprint, hprint, iprint, wprint, eprint, p4print
from Messaging import create_logger, footer, header
from Misc import run_system_cmd, get_release_version, read_legal_release, first_available_file
import CommonHeader
from Misc import utils__script_usage_statistics

__author__ = "Forest Li"
__version__ = get_release_version()


def parse_args():
    parser = argparse.ArgumentParser(description="description here")
    parser.add_argument("-v", metavar="<#>", type=int, default=0, help="verbosity")
    parser.add_argument("-d", metavar="<#>", type=int, default=0, help="debug")
    parser.add_argument("-macro", metavar="<macro>", help="Macro to run on")
    parser.add_argument(
        "-test",
        action="store_true",
        help="Test script and skip submitting usage statistics.",
    )
    parser.add_argument(
        "-p",
        "--project",
        metavar="<project path>",
        help="Project spec, family/project/rel",
        required=True,
    )
    args = parser.parse_args()
    return args


def main(args: argparse.Namespace) -> List[Dict]:
    """runs the spreadsheet model analysis"""
    # Obtain project string
    project_string = args.project
    macro = args.macro
    split_project_string = project_string.split("/")

    if len(split_project_string) != 3:
        fatal_error(
            "Missing element from -proj option, make sure to enter product family, project and version number seperated by '/'."
        )

    # Obtain product family, project path, and release version
    product_family = split_project_string[0]
    project = split_project_string[1]
    release = split_project_string[2]

    dprint(LOW, f"{product_family} / {project} / {release}")

    legal_release = first_available_file([
        f"/remote/cad-rep/projects/{product_family}/{project}/{release}/design/legalRelease.yml",
        f"/remote/cad-rep/projects/{product_family}/{project}/{release}/design/legalRelease.txt",
    ])

    parsed_legal_release = read_legal_release(legal_release)
    release = parsed_legal_release["rel"]
    releaseRoot = parsed_legal_release["p4_release_root"]
    # Find impedance spreadsheet for this project
    spreadsheet_path = find_ibis_impedance_spreadsheet(product_family)

    if spreadsheet_path == NULL_VAL:
        fatal_error("Could not find the impedance report spreadsheet.")

    model_dict = extract_model_names_from_spreadsheet(spreadsheet_path, project_string)

    # Dump models
    for key in model_dict.keys():

        hprint(f"Found {len(model_dict[key])} models for {key}")
        for model in model_dict[key]:
            dprint(SUPER, f"\t{model}")

    if model_dict == NULL_VAL:
        fatal_error("Could not extract model names from the spreadsheet.")

    dprint(INSANE, json.dumps(model_dict))

    # Construct project path to find models
    # if product_family == 'lpddr5x_ddr5_phy':
    #     sub_family = "-ddr5" if ("-ddr5-" in spreadsheet_path) else "lp5x"
    #     project_path = f"//depot/{releaseRoot}"
    # else:

    project_path = f"//depot/{releaseRoot}/ckt/rel"

    missing_models = check_if_models_exist(model_dict, release, project_path, macro)

    report_missing_models(missing_models)

    return missing_models


def find_ibis_impedance_spreadsheet(product_family):
    # Store the sub-product family name, if required
    sub_family = ""

    if product_family == "lpddr5x" or product_family == "ddr5":

        # In this product family the prgoram must ask the user for the sub-product family
        # This is accomplished with the help of the inquirer module
        questions = [
            inquirer.List(
                "sub_project",
                message="Please choose the sub-product family",
                choices=["ddr5", "lp5x"],
            ),
        ]
        answers = inquirer.prompt(questions)
        sub_family = answers["sub_project"]

        supported_impedance_directory = f"//depot/products/lpddr5x_ddr5_phy/{sub_family}/common/qms/templates/IBIS_supported_impedance"

    else:
        supported_impedance_directory = f"//depot/products/{product_family}/common/qms/templates/IBIS_supported_impedance"

    # Debug the directory path
    dprint(HIGH, f"Directory: {supported_impedance_directory}")

    cmd_output, cmd_err, cmd_ret = run_system_cmd(
        f"p4 files {supported_impedance_directory}/...", CommonHeader.VERBOSITY
    )
    cmd_output_array = cmd_output.splitlines()
    dprint(SUPER, cmd_output)

    cmd_output_array = [
        re.sub(r"\#.*", "", i, flags=re.IGNORECASE) for i in cmd_output_array
    ]

    correct_file = NULL_VAL

    if len(cmd_output_array) > 1:

        wprint("More than 1 impedance spreadsheet found.")

        # If more than 1 file is found, then chose the correct file based on the user's input
        questions = [
            inquirer.List(
                "file",
                message="Please choose the correct file for comparison.",
                choices=cmd_output_array,
            ),
        ]
        answers = inquirer.prompt(questions)
        correct_file = answers["file"]

    elif len(cmd_output_array) == 1:
        correct_file = cmd_output_array[0]

    else:
        correct_file = NULL_VAL

    dprint(HIGH, f"{correct_file}")
    if correct_file == NULL_VAL:
        eprint("Could not determine impedance report spreadsheet to download.")

    # Download file to a tempporary location.
    output_file = f"/tmp/{time.time()}_report-spreadhsheet-f{product_family}.xlsx"
    cmd_output, cmd_err, cmd_ret = run_system_cmd(
        f"p4 print -o {output_file} {correct_file}", CommonHeader.VERBOSITY
    )
    dprint(SUPER, f"{cmd_output}")

    if "no such file" in cmd_output or "no such file" in cmd_err:
        dprint(LOW, f"{cmd_err}")
        eprint(
            f"Could not obtain impedance report spreadsheet from {supported_impedance_directory}."
        )

    return output_file


def extract_model_names_from_spreadsheet(spreadsheet_path, project_string):
    model_dict = {}

    try:
        wb = load_workbook(spreadsheet_path)
    except Exception:
        fatal_error("An error occurred while opening the spreadsheet.")

    sheets = wb.sheetnames

    impedance_list = wb[sheets[0]]

    # DDR54 specific checks for v2 PHYs. (For v2 PHYs the data is stored in the second sheet of the ddr54 impedance spreadsheet)
    if "v2" in project_string and "ddr54" in project_string:

        if len(sheets) < 1:
            eprint(
                "More than 2 sheets required for DDR54 impedance report spreadsheet. The second sheet contains the DDR54v2 PHY data."
            )
            return NULL_VAL
        impedance_list = wb[sheets[1]]

    start_positions = []

    for row in impedance_list.iter_rows():
        for cell in row:
            if "START" in str(cell.value):
                curr_start = []
                curr_start.append(cell.row)
                curr_start.append(cell.column)

                start_positions.append(curr_start)

    model_dict = populate_model_dict(impedance_list, start_positions)

    return model_dict


def populate_model_dict(impedance_list, start_positions):
    model_dict = {}

    for impedence_table in start_positions:

        row = impedence_table[0]
        column = impedence_table[1] + 1

        function_offset, model_offset = find_offset(impedance_list, row, column)

        row = impedence_table[0] + 1
        column = impedence_table[1]

        while "STOP" not in str(impedance_list.cell(row=row, column=column).value):

            curr_function = str(
                impedance_list.cell(row=row, column=column + function_offset).value
            ).split(" ")[0]
            if curr_function == "DQ":
                curr_function = "DQ / DQS"

            curr_model = str(
                impedance_list.cell(row=row, column=column + model_offset).value
            ).strip()

            if curr_function != "None" and curr_model != "None":

                if curr_function not in model_dict:

                    new_function = []

                    if "/" in curr_model:
                        split = curr_model.split("/")
                        new_function.append(split[0].strip())
                        new_function.append(split[1].strip())
                    else:
                        new_function.append(curr_model.strip())
                    model_dict[curr_function] = new_function

                else:
                    if "/" in curr_model:
                        split = curr_model.split("/")
                        model_dict[curr_function].append(split[0].strip())
                        model_dict[curr_function].append(split[1].strip())
                    else:
                        model_dict[curr_function].append(curr_model.strip())

            row += 1

    return model_dict


def find_offset(impedance_list, row, column):
    function_offset = 1
    function_stop = False
    model_offset = 1
    model_stop = False

    while "STOP" not in str(impedance_list.cell(row=row, column=column).value):
        if "Function" in str(impedance_list.cell(row=row, column=column).value):
            function_stop = True
        elif not function_stop:
            function_offset += 1

        if "Model name" in str(impedance_list.cell(row=row, column=column).value):
            model_stop = True
        elif not model_stop:
            model_offset += 1

        column += 1

    return function_offset, model_offset


def check_if_models_exist(model_dict, release, project_path, macro):
    # Keep a list of missing models

    missing_ac = []
    missing_dq = []
    missing_ck = []
    missing_alert = []
    missing_cmos = []
    missing_cs = []

    if len(model_dict.keys()) == 0:
        eprint("No models present inside model list for checking.")
        return

    dq_model_array = [
        "dwc_ddrphy_txrxdq_ew",
        "dwc_ddrphy_txrxdqs_ew",
        "dwc_ddrphy_txrxdq_ns",
        "dwc_ddrphy_txrxdqs_ns",
        "dwc_lpddr5xphy_txrxdq_ew",
        "dwc_lpddr5xphy_txrxdq_ns",
        "dwc_lpddr5xphy_txrxdqs_ew",
        "dwc_lpddr5xphy_txrxdqs_ns",
        "dwc_lpddr5xmphy_txrxdqs_ew",
        "dwc_lpddr5xmphy_txrxdqs_ns",
        "dwc_lpddr5xmphy_txrxdq_ew",
        "dwc_lpddr5xmphy_txrxdq_ns",
    ]

    cmos_model_array = [
        "dwc_ddrphy_txrxcmos_ew",
        "dwc_ddrphy_txrxcmos_ns",
        "dwc_lpddr5xphy_txrxcmos_ew",
        "dwc_lpddr5xphy_txrxcmos_ns",
        "dwc_lpddr5xmphy_txrxcmos_ew",
        "dwc_lpddr5xmphy_txrxcmos_ns",
    ]

    cs_model_array = [
        "dwc_ddrphy_txrxcs_ew",
        "dwc_ddrphy_txrxcs_ns",
        "dwc_lpddr5xphy_txrxcs_ew",
        "dwc_lpddr5xphy_txrxcs_ns",
        "dwc_lpddr5xmphy_txrxcs_ew",
        "dwc_lpddr5xmphy_txrxcs_ns",
    ]

    ac_model_array = [
        "dwc_ddrphy_txrxac_ew",
        "dwc_ddrphy_txrxac_ns",
        "dwc_ddrphy_txrxca_ew",
        "dwc_ddrphy_txrxca_ns",
        "dwc_lpddr5xphy_txrxac_ns",
        "dwc_lpddr5xphy_txrxac_ew",
        "dwc_lpddr5xmphy_txrxac_ew",
        "dwc_lpddr5xmphy_txrxac_ns",
    ]

    # Algorithm
    # For each TXRX type
    #   - Check whether an appropraite summary file exists
    #   - If that file exists, then verify that its models exist inside the excel spreadsheet
    if macro is None:
        (
            missing_ac,
            missing_dq,
            missing_ck,
            missing_alert,
            missing_cmos,
            missing_cs,
        ) = populate_missing_models(
            project_path,
            release,
            model_dict,
            dq_model_array,
            cmos_model_array,
            cs_model_array,
            ac_model_array,
        )

    else:
        (
            missing_ac,
            missing_dq,
            missing_ck,
            missing_alert,
            missing_cmos,
            missing_cs,
        ) = populate_missing_models_macro(
            macro,
            project_path,
            release,
            model_dict,
            dq_model_array,
            cmos_model_array,
            cs_model_array,
            ac_model_array,
        )

    all_missing_models = {}

    if missing_ac:
        all_missing_models["AC"] = missing_ac

    if missing_dq:
        all_missing_models["DQ / DQS"] = missing_dq

    if missing_ck:
        all_missing_models["CK"] = missing_ck

    if missing_alert:
        all_missing_models["ALERT"] = missing_alert

    if missing_cmos:
        all_missing_models["CMOS"] = missing_cmos

    if missing_cs:
        all_missing_models["CS"] = missing_cs

    return all_missing_models


def populate_missing_models_macro(
    macro,
    project_path,
    release,
    model_dict,
    dq_model_array,
    cmos_model_array,
    cs_model_array,
    ac_model_array,
):
    missing_ac = []
    missing_dq = []
    missing_ck = []
    missing_alert = []
    missing_cmos = []
    missing_cs = []
    found = FALSE

    for key in model_dict:

        if key == "DQ / DQS" and re.search("dq", macro):
            missing_dq = check_model_list(
                project_path,
                release,
                dq_model_array,
                model_dict,
                key,
                missing_dq,
                macro,
            )
            found = TRUE
            break

        elif key == "CMOS" and re.search("cmos", macro):
            missing_cmos = check_model_list(
                project_path,
                release,
                cmos_model_array,
                model_dict,
                key,
                missing_cmos,
                macro,
            )
            found = TRUE
            break

        elif key == "CS" and re.search("cs", macro):
            missing_cs = check_model_list(
                project_path,
                release,
                cs_model_array,
                model_dict,
                key,
                missing_cs,
                macro,
            )
            found = TRUE
            break

        elif key == "AC" and re.search("ac", macro):
            missing_ac = check_model_list(
                project_path,
                release,
                ac_model_array,
                model_dict,
                key,
                missing_ac,
                macro,
            )
            missing_alert = missing_ac
            missing_ck = missing_ac
            found = TRUE
            break

    if found == FALSE:
        fatal_error(f"There is no {macro} macro in this project!")

    return missing_ac, missing_dq, missing_ck, missing_alert, missing_cmos, missing_cs


def populate_missing_models(
    project_path,
    release,
    model_dict,
    dq_model_array,
    cmos_model_array,
    cs_model_array,
    ac_model_array,
):
    missing_ac = []
    missing_dq = []
    missing_ck = []
    missing_alert = []
    missing_cmos = []
    missing_cs = []
    macro = None

    for key in model_dict:

        if key == "DQ / DQS":
            missing_dq = check_model_list(
                project_path,
                release,
                dq_model_array,
                model_dict,
                key,
                missing_dq,
                macro,
            )

        elif key == "CMOS":
            missing_cmos = check_model_list(
                project_path,
                release,
                cmos_model_array,
                model_dict,
                key,
                missing_cmos,
                macro,
            )

        elif key == "CS":
            missing_cs = check_model_list(
                project_path,
                release,
                cs_model_array,
                model_dict,
                key,
                missing_cs,
                macro,
            )

        elif key == "AC":
            missing_ac = check_model_list(
                project_path,
                release,
                ac_model_array,
                model_dict,
                key,
                missing_ac,
                macro,
            )
            missing_alert = missing_ac
            missing_ck = missing_ac

    return missing_ac, missing_dq, missing_ck, missing_alert, missing_cmos, missing_cs


def check_model_list(
    project_path, release, model_list, model_dict, key, missing_list, macro
):
    if macro is not None:
        wprint(f"Macro is defined as {macro}, only looking for those models")
        summary_file_path = f'{project_path}/{macro}/{release}/macro/ibis/"*summary*"'
        result = check_model(summary_file_path, model_dict[key], missing_list)
        if result == NULL_VAL:
            pass
        else:
            missing_list = result

    else:
        for curr_model in model_list:

            summary_file_path = (
                f'{project_path}/{curr_model}/{release}/macro/ibis/"*summary*"'
            )
            result = check_model(summary_file_path, model_dict[key], missing_list)

            if result == NULL_VAL:
                pass
            else:
                missing_list = result

    return missing_list


def check_model(summary_file_path, models, missing_models):
    # TODO: Replace with da_p4_files from P4Utils.py
    cmd_output, cmd_err, cmd_ret = run_system_cmd(
        f"p4 files -e {summary_file_path}", CommonHeader.VERBOSITY
    )

    if "no such file" in cmd_output or "no such file" in cmd_err:
        wprint(f"Could not find: {summary_file_path}")
        return NULL_VAL

    else:

        file_name_array = cmd_output.splitlines()

        if len(file_name_array) == 0:
            wprint(f"Could not find summary files at {summary_file_path}")
            return NULL_VAL

        file_name_array = [
            re.sub(r"\#.*", "", i, flags=re.IGNORECASE) for i in file_name_array
        ]

        # Always get the first summary file. (CAVEAT: Multiple summary files may exist)
        file_name = file_name_array[0]

        iprint(f"File found: {file_name}")
        result, cmd_err, cmd_ret = run_system_cmd(
            f"p4 print -q {file_name}", CommonHeader.VERBOSITY
        )
        dprint(INSANE, result)

        for model in models:
            if model not in result:
                missing_models.append(model)

    # Gather unique items only
    myset = set(missing_models)
    missing_models = list(myset)

    return missing_models


def report_missing_models(missing_models):
    if not missing_models:
        p4print("PASS\n")
        return
    for model_name, models in missing_models.items():
        hprint(
            f"The following {len(models)} {model_name} function models were not found in the ibis summary file(s):"
        )
        for model in model_name:
            iprint(f"\t{model}")


if __name__ == "__main__":

    cmd_line_args = parse_args()
    filename = os.getcwd() + "/" + os.path.basename(__file__) + ".log"
    logger = create_logger(filename)  # Create log file

    # Register exit function
    atexit.register(footer)

    # Initalise shared variables and run main
    CommonHeader.init(cmd_line_args, __author__, __version__)

    header()
    main(cmd_line_args)
    iprint(f"Log file: {filename}")

    if cmd_line_args.test:
        pass
    else:
        utils__script_usage_statistics("ibis_impedance_spreadsheet", "2022ww22")
