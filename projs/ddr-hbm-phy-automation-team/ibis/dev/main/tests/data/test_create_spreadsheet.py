#!/depot/Python/Python-3.8.0/bin/python
###############################################################################
# Name    : test_create_spreadsheet.py
# Author  : Raneem Khalil
# Date     : 2022-07-13 15:34:41
# Purpose : Testcases for test_ibis_correlation_report.py
###############################################################################
# nolint main
# nolint utils__script_usage_statistics
__author__ = 'Raneem Khalil'
__version__ = '2022ww28'

import pathlib
import sys
import openpyxl
import random
from openpyxl.styles import Alignment


# ---------------------------------- #
bindir = str(pathlib.Path(__file__).resolve().parent)
sys.path.append(bindir + '/../../lib/python/Util')
sys.path.append(bindir + '/../../lib/python')
# ---------------------------------- #


# The sheet including IBIS information exists
def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0, title="Output")
    return wb, ws


# The sheet including IBIS information does not exist
def create_workbook_incorrect_sheet_name():
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title="Other_output")
    wb.save('correlationReportIncorrectSheetName.xlsx')


# File is not saved in the correct format
def save_html(ws, wb):
    ws['E1'].value = 'IBIS'
    ws.merge_cells('E1:H1')
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E2'].value = 'Case'
    ws['F2'].value = 'PU resis'
    ws['G2'].value = 'PD resis'
    ws['H2'].value = 'ODT resis'

    num_drv = random.randint(20, 85)
    min_num = (num_drv - num_drv / 10) * 100
    max_num = (num_drv + num_drv / 10) * 100

    num_rcv = random.randint(20, 85)
    min_num_odt = (num_rcv - num_rcv / 10) * 100
    max_num_odt = (num_rcv + num_rcv / 10) * 100

    cases = ['drv max', str(num_drv), 'drv min', str(num_drv), 'drv typ', str(num_drv), 'rcv max', 'off', 'rcv min', str(num_rcv), 'rcv typ', str(num_rcv)]
    j = 3
    for i in range(len(cases)):
        cell = ws.cell(row=j, column=5)
        cell.value = cases[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 1

    num_in_range_pu_max = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pu_min = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pu_typ = round((float(random.uniform(min_num, max_num)) / 100), 4)
    pu = [num_in_range_pu_max, num_in_range_pu_min, num_in_range_pu_typ]
    j = 4
    for i in range(len(pu)):
        cell = ws.cell(row=j, column=6)
        cell.value = pu[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_in_range_pd_max = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pd_min = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pd_typ = round((float(random.uniform(min_num, max_num)) / 100), 4)
    pd = [num_in_range_pd_max, num_in_range_pd_min, num_in_range_pd_typ]
    j = 4
    for i in range(len(pd)):
        cell = ws.cell(row=j, column=7)
        cell.value = pd[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_off = random.randrange(10000, 100000)
    num_in_range_odt_min = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    num_in_range_odt_typ = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    odt = [num_off, num_in_range_odt_min, num_in_range_odt_typ]
    j = 10
    for i in range(len(odt)):
        cell = ws.cell(row=j, column=8)
        cell.value = odt[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    wb.save('correlationReport.html')


# Everything is in the correct format
def write_to_spreadsheet_correct_format(ws, wb):
    ws['E1'].value = 'IBIS'
    ws.merge_cells('E1:H1')
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E2'].value = 'Case'
    ws['F2'].value = 'PU resis'
    ws['G2'].value = 'PD resis'
    ws['H2'].value = 'ODT resis'

    num_drv = random.randint(20, 85)
    min_num = (num_drv - num_drv / 10) * 100
    max_num = (num_drv + num_drv / 10) * 100

    num_rcv = random.randint(20, 85)
    min_num_odt = (num_rcv - num_rcv / 10) * 100
    max_num_odt = (num_rcv + num_rcv / 10) * 100

    cases = ['drv max', str(num_drv), 'drv min', str(num_drv), 'drv typ', str(num_drv), 'rcv max', 'off', 'rcv min', str(num_rcv), 'rcv typ', str(num_rcv)]
    j = 3
    for i in range(len(cases)):
        cell = ws.cell(row=j, column=5)
        cell.value = cases[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 1

    num_in_range_pu_max = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pu_min = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pu_typ = round((float(random.uniform(min_num, max_num)) / 100), 4)
    pu = [num_in_range_pu_max, num_in_range_pu_min, num_in_range_pu_typ]
    j = 4
    for i in range(len(pu)):
        cell = ws.cell(row=j, column=6)
        cell.value = pu[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_in_range_pd_max = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pd_min = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pd_typ = round((float(random.uniform(min_num, max_num)) / 100), 4)
    pd = [num_in_range_pd_max, num_in_range_pd_min, num_in_range_pd_typ]
    j = 4
    for i in range(len(pd)):
        cell = ws.cell(row=j, column=7)
        cell.value = pd[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_off = random.randrange(10000, 100000)
    num_in_range_odt_min = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    num_in_range_odt_typ = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    odt = [num_off, num_in_range_odt_min, num_in_range_odt_typ]
    j = 10
    for i in range(len(odt)):
        cell = ws.cell(row=j, column=8)
        cell.value = odt[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    wb.save('correlationReport.xlsx')


# Correct format but there are failed models
def write_to_spreadsheet_failed_models(ws, wb):
    ws['E1'].value = 'IBIS'
    ws.merge_cells('E1:H1')
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E2'].value = 'Case'
    ws['F2'].value = 'PU resis'
    ws['G2'].value = 'PD resis'
    ws['H2'].value = 'ODT resis'

    num_drv = random.randint(20, 85)
    min_num = (num_drv - num_drv / 10) * 100
    max_num = (num_drv + num_drv / 10) * 100

    num_rcv = random.randint(20, 85)
    min_num_odt = (num_rcv - num_rcv / 10) * 100
    max_num_odt = (num_rcv + num_rcv / 10) * 100

    cases = ['drv max', str(num_drv), 'drv min', str(num_drv), 'drv typ', str(num_drv), 'rcv max', 'off', 'rcv min', str(num_rcv), 'rcv typ', str(num_rcv)]
    j = 3
    for i in range(len(cases)):
        cell = ws.cell(row=j, column=5)
        cell.value = cases[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 1

    num_out_of_range_pu_max = round((float(random.uniform(50, min_num - 9.9999)) / 100), 4)
    num_in_range_pu_min = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_out_of_range_pu_typ = round((float(random.uniform(50, min_num - 9.9999)) / 100), 4)
    pu = [num_out_of_range_pu_max, num_in_range_pu_min, num_out_of_range_pu_typ]
    j = 4
    for i in range(len(pu)):
        cell = ws.cell(row=j, column=6)
        cell.value = pu[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_in_range_pd_max = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_out_of_range_pd_min = round((float(random.uniform(max_num + 0.0001, 100)) / 100), 4)
    num_in_range_pd_typ = round((float(random.uniform(min_num, max_num)) / 100), 4)
    pd = [num_in_range_pd_max, num_out_of_range_pd_min, num_in_range_pd_typ]
    j = 4
    for i in range(len(pd)):
        cell = ws.cell(row=j, column=7)
        cell.value = pd[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_off = random.randint(100, 999)
    num_in_range_odt_min = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    num_out_of_range_odt_typ = round((float(random.uniform(max_num_odt + 0.0001, 100)) / 100), 4)
    odt = [num_off, num_in_range_odt_min, num_out_of_range_odt_typ]
    j = 10
    for i in range(len(odt)):
        cell = ws.cell(row=j, column=8)
        cell.value = odt[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    wb.save('correlationReportFailedModels.xlsx')


# Spreadsheet does not include IBIS Columns
def write_to_spreadsheet_no_ibis_columns(ws, wb):
    ws['E1'].value = 'HSPICE'
    ws.merge_cells('E1:H1')
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E2'].value = 'Case'
    ws['F2'].value = 'PU resis'
    ws['G2'].value = 'PD resis'
    ws['H2'].value = 'ODT resis'

    num_drv = random.randint(20, 85)
    min_num = (num_drv - num_drv / 10) * 100
    max_num = (num_drv + num_drv / 10) * 100

    num_rcv = random.randint(20, 85)
    min_num_odt = (num_rcv - num_rcv / 10) * 100
    max_num_odt = (num_rcv + num_rcv / 10) * 100

    cases = ['drv max', str(num_drv), 'drv min', str(num_drv), 'drv typ', str(num_drv), 'rcv max', 'off', 'rcv min', str(num_rcv), 'rcv typ', str(num_rcv)]
    j = 3
    for i in range(len(cases)):
        cell = ws.cell(row=j, column=5)
        cell.value = cases[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 1

    num_in_range_pu_max = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pu_min = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pu_typ = round((float(random.uniform(min_num, max_num)) / 100), 4)
    pu = [num_in_range_pu_max, num_in_range_pu_min, num_in_range_pu_typ]
    j = 4
    for i in range(len(pu)):
        cell = ws.cell(row=j, column=6)
        cell.value = pu[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_in_range_pd_max = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pd_min = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pd_typ = round((float(random.uniform(min_num, max_num)) / 100), 4)
    pd = [num_in_range_pd_max, num_in_range_pd_min, num_in_range_pd_typ]
    j = 4
    for i in range(len(pd)):
        cell = ws.cell(row=j, column=7)
        cell.value = pd[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_off = random.randint(10000, 100000)
    num_in_range_odt_min = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    num_in_range_odt_typ = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    odt = [num_off, num_in_range_odt_min, num_in_range_odt_typ]
    j = 10
    for i in range(len(odt)):
        cell = ws.cell(row=j, column=8)
        cell.value = odt[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    wb.save('correlationReportNoIbisColumns.xlsx')


# One or more of the IBIS columns is missing
def write_to_spreadsheet_missing_columns(ws, wb):
    ws['E1'].value = 'IBIS'
    ws.merge_cells('E1:H1')
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E2'].value = 'Case'
    ws['F2'].value = 'PU resis'
    ws['H2'].value = 'ODT resis'

    num_drv = random.randint(20, 85)
    min_num = (num_drv - num_drv / 10) * 100
    max_num = (num_drv + num_drv / 10) * 100

    num_rcv = random.randint(20, 85)
    min_num_odt = (num_rcv - num_rcv / 10) * 100
    max_num_odt = (num_rcv + num_rcv / 10) * 100

    cases = ['drv max', str(num_drv), 'drv min', str(num_drv), 'drv typ', str(num_drv), 'rcv max', 'off', 'rcv min', str(num_rcv), 'rcv typ', str(num_rcv)]
    j = 3
    for i in range(len(cases)):
        cell = ws.cell(row=j, column=5)
        cell.value = cases[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 1

    num_in_range_pu_max = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pu_min = round((float(random.uniform(min_num, max_num)) / 100), 4)
    num_in_range_pu_typ = round((float(random.uniform(min_num, max_num)) / 100), 4)
    pu = [num_in_range_pu_max, num_in_range_pu_min, num_in_range_pu_typ]
    j = 4
    for i in range(len(pu)):
        cell = ws.cell(row=j, column=6)
        cell.value = pu[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    num_off = random.randint(10000, 100000)
    num_in_range_odt_min = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    num_in_range_odt_typ = round((float(random.uniform(min_num_odt, max_num_odt)) / 100), 4)
    odt = [num_off, num_in_range_odt_min, num_in_range_odt_typ]
    j = 10
    for i in range(len(odt)):
        cell = ws.cell(row=j, column=8)
        cell.value = odt[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        j += 2

    wb.save('correlationReportMissingColumns.xlsx')
