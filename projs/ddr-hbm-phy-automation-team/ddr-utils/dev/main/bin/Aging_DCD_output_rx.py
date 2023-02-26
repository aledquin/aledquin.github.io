#!/depot/Python/Python-3.8.0/bin/python

"""

Aging_DCD_output_rx.py

Find the desired corners from MPW.ALL/(ProjectSheet) and append
to the end of the tables in Aging_DCD.../(OutputSheet) for:

Parameters:

1. DICLk and FClk (rxclk_outdi & rxclk_dqrxflop)
2. High(Temp) and Low(Temp)
3  orient (macro orientation)
3. Boost and/or Non-Boost
4. MC and DI Margin


Instructions on usage:

1. Download MPW & Aging_DCD xlsx sheets in current path
2. Edit the parameters in the config file (multiple projects supported)
3. Run using ..python Aging_DCD_output_rx.py -c config_aging.ini
4. This script will append the data to the end of the table defined in the
   Aging_DCD excel file
"""

import os
import re
import argparse
import pathlib
import atexit

import math
import sys
import configparser

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------- #
bindir = str(pathlib.Path(__file__).resolve().parent)
# Add path to library that may be symbolically linked.
sys.path.append(bindir + "/../lib/Util")
# Add path to sharedlib's Python Utilities directory.
sys.path.append(bindir + "/../lib/python/Util")
# Add path to Aging package directory.
sys.path.append(bindir + "/../lib/python")
# ---------------------------------- #

from CommonHeader import LOW, MEDIUM, HIGH
from Messaging import iprint, hprint, fatal_error, nprint, dprint
from Misc import utils__script_usage_statistics, get_release_version
from Messaging import create_logger, footer, header
import CommonHeader

import aging_pkg

__author__ = "Karen Wan"
__tool_name__ = "Aging_DCD_output_rx"


# -----------------------------------------------------------------------------
# --------------------------------MAIN PROGRAM---------------------------------
# -----------------------------------------------------------------------------
def main(args):
    global param_percentage
    global automotive_proj
    global temp_high
    global temp_low
    global mc_data_b
    global mc_data_nb
    inputConfig = args

    aging_pkg.check_config_exists(inputConfig.cfile)

    config = configparser.ConfigParser()
    config.read(inputConfig.cfile)

    # hprint(f'\n{"LOG OUTPUT".center(50, "-")}')

    for project in config.sections():

        MPWinput_wb = config[project]["MPWinput_wb"]
        MPWinput_ws = config[project]["MPWinput_ws"]
        DCDoutput_wb = config[project]["DCDoutput_wb"]
        DCDoutput_ws = config[project]["DCDoutput_ws"]
        orient = config[project]["orient"]
        high_nb = config[project]["high_nb"]
        low_nb = config[project]["low_nb"]
        high_b = config[project]["high_b"]
        low_b = config[project]["low_b"]
        projectName = config[project]["projectName"]
        di_margin = float(config[project]["di_margin"])
        param_percentage = float(config[project]["param_percentage"])

        mc_data_b = config[project]["mc_data_b"]
        mc_data_nb = config[project]["mc_data_nb"]
        temp_low = config[project]["temp_low"]
        temp_high = config[project]["temp_high"]
        automotive_project = config[project]["automotive_project"]

        # ppmin_data_b = config[project]['ppmin_data_b']
        # ppmin_data_nb = config[project]['ppmin_data_nb']

        automotive_proj = False
        if automotive_project.lower() in ["true", "1", "t", "y", "yes"]:
            automotive_proj = True

        # #####################################################################
        if mc_data_b == "NA" and mc_data_nb == "NA":
            fatal_error(
                (
                    "ERROR:\tMonte Carlo boost/non-boost not defined -"
                    " Check config. Exiting..."
                )
            )

        iprint(f"The MC Data used for project: {projectName}")
        # df_b = pd.read_excel(mc_data_b)
        # df_nb = pd.read_excel(mc_data_nb)
        # matches_f = ["fclk", "std"]
        matches_f = [
            "rxclk_dqrxflop_pn",
            "std",
        ]  # Only use the monte results from negative pulse
        # matches_di = [ "diclk" , "std" ]
        matches_di = [
            "rxclk_outdi_pn",
            "std",
        ]  # Only use the monte results from negative pulse

        # --Boost
        if mc_data_b != "NA":
            df_b = pd.read_excel(mc_data_b, engine="openpyxl")
            mc_fclk_low_boost = mc_final(df_b, temp_low, matches_f)
            nprint(f"\tmc_fclk_low_boost: {str(mc_fclk_low_boost)}")
            mc_fclk_high_boost = mc_final(df_b, temp_high, matches_f)
            nprint(f"\tmc_fclk_high_boost: {str(mc_fclk_high_boost)}")
            mc_diclk_low_boost = mc_final(df_b, temp_low, matches_di)
            nprint(f"\tmc_diclk_low_boost: {str(mc_diclk_low_boost)}")
            mc_diclk_high_boost = mc_final(df_b, temp_high, matches_di)
            nprint(f"\tmc_diclk_high_boost: {str(mc_diclk_high_boost)}")
        else:
            mc_fclk_low_boost = 0
            mc_fclk_high_boost = 0
            mc_diclk_low_boost = 0
            mc_diclk_high_boost = 0

        # --Non Boost
        if mc_data_nb != "NA":
            df_nb = pd.read_excel(mc_data_nb, engine="openpyxl")
            mc_fclk_low_nonboost = mc_final(df_nb, temp_low, matches_f)
            nprint(f"\tmc_fclk_low_nonboost: {str(mc_fclk_low_nonboost)}")
            mc_fclk_high_nonboost = mc_final(df_nb, temp_high, matches_f)
            nprint(f"\tmc_fclk_high_nonboost: {str(mc_fclk_high_nonboost)}")
            mc_diclk_low_nonboost = mc_final(df_nb, temp_low, matches_di)
            nprint(f"\tmc_diclk_low_nonboost: {str(mc_diclk_low_nonboost)}")
            mc_diclk_high_nonboost = mc_final(df_nb, temp_high, matches_di)
            nprint(f"\tmc_diclk_high_nonboost: {str(mc_diclk_high_nonboost)}")
        else:
            mc_fclk_low_nonboost = 0
            mc_fclk_high_nonboost = 0
            mc_diclk_low_nonboost = 0
            mc_diclk_high_nonboost = 0

        # #####################################################################
        agingDcdOutput(
            MPWinput_wb,
            MPWinput_ws,
            DCDoutput_wb,
            DCDoutput_ws,
            orient,
            high_nb,
            low_nb,
            high_b,
            low_b,
            projectName,
            di_margin,
            mc_diclk_high_boost,
            mc_diclk_low_boost,
            mc_diclk_high_nonboost,
            mc_diclk_low_nonboost,
            mc_fclk_high_boost,
            mc_fclk_low_boost,
            mc_fclk_high_nonboost,
            mc_fclk_low_nonboost,
        )

    hprint(f'{"COMPLETED".center(50, "-")}')


# -----------------------------------------------------------------------------
# ------------------------------FUNCTIONS DEFINED------------------------------
# -----------------------------------------------------------------------------


# function returns the command line argument
def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Find the desired corners from MPW.ALL/(ProjectSheet) and"
            " append to the end of the tables in Aging_DCD.../"
            "(OutputSheet) for:\n\nParameters:\n1. DICLk and FClk "
            "(rxclk_outdi & rxclk_dqrxflop)\n2. High(Temp) and "
            "Low(Temp)\n3  orient (macro orientation)\n3. "
            "Boost and/or Non-Boost\n4. MC and DI Margin\n\n"
            "Instructions on usage:\n\n1. Download MPW & "
            "Aging_DCD xlsx sheets in current path\n2. "
            "Edit the parameters in the config file "
            "(multiple projects supported)\n3. Run using ..python "
            "Aging_DCD_output_pclk.py -c config_aging.ini\n4. "
            "This script will append the data to the end of the "
            "table defined in the Aging_DCD excel file"
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "-v", metavar="<#>", type=int, default=0, help="verbosity"
    )
    parser.add_argument("-d", metavar="<#>", type=int, default=0, help="debug")
    # -------------------------------------
    parser.add_argument(
        "-c", "--cfile", help="REQUIRED. Config file path", required=True
    )
    # -------------------------------------
    args = parser.parse_args()
    return args


# ########################## FUNCTIONS FOR MC OUTPUT ##########################
def mc_out(array, temp, df):

    dataFrameArr = ["mos", "temp"] + array
    clk_df = df.filter(dataFrameArr, axis=1)
    # clk_df_x = clk_df[(clk_df['mos'] == 'mos_ss') & (clk_df['temp'] == temp)]
    clk_df_x = clk_df[(clk_df["temp"] == temp)]
    clk_row_x = clk_df_x.iloc[0]
    clk_row_x[array] = clk_row_x[array].apply(pd.to_numeric)

    max_fromdf = pd.DataFrame(clk_row_x[array])
    mc_number = float(max_fromdf.max(axis=0).iloc[0])

    if automotive_proj:
        # monte value used for automotive project changed from std*5 to std*3
        returnvalue = round(mc_number * 3, 2)

    else:
        returnvalue = round(mc_number * 3, 2)

    return returnvalue


def mc_final(df, temp, matches):

    array_clk = aging_pkg.columnArrOut(df, matches)
    mc_clk = mc_out(array_clk, temp, df)

    return mc_clk


# ######################## Functions for PPMIN OUTPUT #########################


def ppmin(excel_data, temp, clk, period, pwd):

    df = pd.read_excel(excel_data, engine="openpyxl")

    # ########################### PPMIN CALCULATION ###########################

    if automotive_proj:
        fcolumn_name = "rxclk_dqrxflop_ppmin_pw_med-target (ps)"
        dicolumn_name = "rxclk_outdi_ppmin_pw_med-target (ps)"
    else:
        fcolumn_name = "rxclk_dqrxflop_ppmin_pw_med-3s (ps)"
        dicolumn_name = "rxclk_outdi_ppmin_pw_med-3s (ps)"

    if clk == "di":
        column_name = dicolumn_name

    else:
        column_name = fcolumn_name

    df_ppmin = df[(df["temp"] == temp)]

    # ################################## CLK ##################################

    clk_row_ppmin = df_ppmin[column_name].astype(float).idxmin()
    clk_ppmin_final = df.iloc[clk_row_ppmin, df.columns.get_loc(column_name)]
    clk_ppmin_process = df.iloc[clk_row_ppmin, df.columns.get_loc("mos")]

    nprint(
        (
            f"\t{str(column_name)}: {str(clk_ppmin_final)}"
            f"{str(clk_ppmin_process)} {str(temp)}"
        )
    )
    a = (period / 2) - pwd
    final_margin = clk_ppmin_final - a

    final_array = [
        str(round(clk_ppmin_final, 2)),
        str(clk_ppmin_process),
        str(round(final_margin, 2)),
    ]

    # Add in the log file here (Then take off the array part!)

    return final_array


# #############################################################################
def writer_f(
    data_f,
    mc_high,
    mc_low,
    di_f,
    col_f,
    row_f,
    currentrow_f,
    projectName,
    ws,
    boost_status,
    clk_status,
):
    """Writer function inputs data into the excel table"""
    aging_pkg.check_corners(data_f)

    for i, j in zip(range(1, 5), range(1, 5)):

        projectnameLow = projectName + " Low"
        projectnameHigh = projectName + " High"

        if j > 2:
            projectStart = projectnameHigh
            # mc_f = mc_high
        else:
            projectStart = projectnameLow
            # mc_f = mc_low

        if (j % 2) == 0:
            mc_f = mc_low
            temp = temp_low
        else:
            mc_f = mc_high
            temp = temp_high

        try:
            f = float(data_f[i - 1].iloc[0]["period"]) * 1000
        except IndexError as error:
            fatal_error(
                f"ERROR:\t{repr(error)}. Missing corner data. Exiting...",
            )

        # param_percentage = 2.5 by default
        ix = param_percentage * (f / 100)
        iy = math.sqrt(pow(mc_f, 2) + pow(ix, 2))
        ir = float(ix + iy)

        h = float(data_f[i - 1].iloc[0]["tpwda"]) * 1000
        jr = h - ir
        k = (f / 2) - ir
        l_ = ((h - ir) * 100) / f
        m = l_ - di_f

        # Project
        ws[(col_f) + str(row_f + i + currentrow_f)] = projectStart
        # Period of operation, ps
        ws[chr(ord(col_f) + 1) + str(row_f + i + currentrow_f)] = f
        # Corner
        ws[chr(ord(col_f) + 2) + str(row_f + i + currentrow_f)] = data_f[
            i - 1
        ].iloc[0]["pvt"]
        # Total pulse width degradation allowed, ps
        ws[chr(ord(col_f) + 3) + str(row_f + i + currentrow_f)] = h
        # PLL 2.5% + RSS( PLL 2.5% + Clocktree monte), ps
        ws[chr(ord(col_f) + 4) + str(row_f + i + currentrow_f)] = round(ir, 2)
        # Aging PWD Budget limit, ps
        ws[chr(ord(col_f) + 5) + str(row_f + i + currentrow_f)] = round(jr, 2)
        # Minimum PW at Input, ps
        ws[chr(ord(col_f) + 6) + str(row_f + i + currentrow_f)] = round(k, 2)
        # % DCD budget for Aging clktree
        ws[chr(ord(col_f) + 7) + str(row_f + i + currentrow_f)] = round(l_, 2)
        # % DCD budget for Aging clktree+lcdl( add 2% DI margin)
        ws[chr(ord(col_f) + 8) + str(row_f + i + currentrow_f)] = round(m, 2)

        # --Add in the ppmin margins:
        if boost_status == "b":
            excel_data = mc_data_b
        else:
            excel_data = mc_data_nb

        clk = clk_status

        if (mc_data_b != "NA" or mc_data_nb != "NA") and j > 2:

            ppmin_arr = ppmin(excel_data, temp, clk, f, h)
            # Positive Pulse Margin, ps
            ws[chr(ord(col_f) + 9) + str(row_f + i + currentrow_f)] = float(
                ppmin_arr[2]
            )

        j = j + 1


def agingDcdOutput(
    MPWinput_wb,
    MPWinput_ws,
    DCDoutput_wb,
    DCDoutput_ws,
    orient,
    high_nb,
    low_nb,
    high_b,
    low_b,
    projectName,
    di_margin,
    mc_diclk_high_boost,
    mc_diclk_low_boost,
    mc_diclk_high_nonboost,
    mc_diclk_low_nonboost,
    mc_fclk_high_boost,
    mc_fclk_low_boost,
    mc_fclk_high_nonboost,
    mc_fclk_low_nonboost,
):
    """
    The main function that the program runs in a loop per project
    defined in the config file
    """
    if "NA" in {MPWinput_wb, MPWinput_ws, DCDoutput_wb, DCDoutput_ws}:
        fatal_error(
            (
                "ERROR:\tI/O Files or Sheets not defined - "
                "Check MPW/DCD file config. Exiting..."
            )
        )

    dprint(LOW, "Checking high_nb and low_nb config entries...")
    aging_pkg.check_high_low(high_nb, low_nb)
    dprint(LOW, "Checking high_b and low_b config entries...")
    aging_pkg.check_high_low(high_b, low_b)

    df = pd.read_excel(MPWinput_wb, sheet_name=MPWinput_ws, engine="openpyxl")
    df.drop(df.columns[[1, 8]], axis=1, inplace=True)
    dprint(HIGH, df)
    df.columns = [
        "macro",
        "clock",
        "wpslack",
        "period",
        "halfclk",
        "pvt",
        "pulse",
    ]

    # SANITY CHECKS

    df.dropna(inplace=True)
    df.drop(df.index[df["halfclk"] == "NA(uncert)"], inplace=True)

    # df.drop(df.index[df.halfclk.str.contains(r'[a-zA-Z]'), inplace = True)
    # df.drop(df.index[df.wpslack.str.contains(r'[a-zA-Z]'), inplace = True)

    # df.drop(df[df.halfclk.str.contains(r'[a-zA-Z]')].index, inplace = True)
    # df.drop(df[df.wpslack.str.contains(r'[a-zA-Z]')].index, inplace = True)

    # df[df.halfclk.apply(lambda x: x.isnumeric())]
    # df[df.wpslack.apply(lambda x: x.isnumeric())]

    df["tpwda"] = df["wpslack"] - df["halfclk"]
    df.drop(["wpslack", "halfclk"], axis=1, inplace=True)
    df = aging_pkg.orientation(orient, df)
    # -------------------------Filter desired corners--------------------------

    # Desired clock source domain names
    clk_names = (
        "RxStrobeCClk_LCDL_SE|RxStrobeTClk_LCDL_SE"
        "|RxDQSEnSync_Clk|RxStrobeCClk|RxStrobeTClk"
    )
    clk_src = df["clock"].str.contains(clk_names)

    # Non - Boost

    if high_nb != "NA" and low_nb != "NA":

        # --diclk
        high_nb_len = len(high_nb)
        low_nb_len = len(low_nb)
        df_dcknb125low = df.loc[
            (clk_src)
            & (df["pvt"].str[0:high_nb_len] == high_nb)
            & (df["pulse"] == "(low)")
        ]
        dprint(HIGH, df_dcknb125low)
        df_dcknbn40low = df.loc[
            (clk_src)
            & (df["pvt"].str[0:low_nb_len] == low_nb)
            & (df["pulse"] == "(low)")
        ]
        dprint(HIGH, df_dcknbn40low)
        df_dcknb125hi = df.loc[
            (clk_src)
            & (df["pvt"].str[0:high_nb_len] == high_nb)
            & (df["pulse"] == "(high)")
        ]
        dprint(HIGH, df_dcknb125hi)
        df_dcknbn40hi = df.loc[
            (clk_src)
            & (df["pvt"].str[0:low_nb_len] == low_nb)
            & (df["pulse"] == "(high)")
        ]
        dprint(HIGH, df_dcknbn40hi)

        df_dcknb125lowst = df_dcknb125low[
            df_dcknb125low.tpwda == df_dcknb125low.tpwda.min()
        ].head(1)
        dprint(HIGH, df_dcknb125lowst)
        df_dcknbn40lowst = df_dcknbn40low[
            df_dcknbn40low.tpwda == df_dcknbn40low.tpwda.min()
        ].head(1)
        dprint(HIGH, df_dcknbn40lowst)
        df_dcknb125hist = df_dcknb125hi[
            df_dcknb125hi.tpwda == df_dcknb125hi.tpwda.min()
        ].head(1)
        dprint(HIGH, df_dcknb125hist)
        df_dcknbn40hist = df_dcknbn40hi[
            df_dcknbn40hi.tpwda == df_dcknbn40hi.tpwda.min()
        ].head(1)
        dprint(HIGH, f"df_pcknbn40hist: {df_dcknbn40hist}")

        diclk_dfnblist = [
            df_dcknb125lowst,
            df_dcknbn40lowst,
            df_dcknb125hist,
            df_dcknbn40hist,
        ]
        dprint(HIGH, f"diclk_dfnblist: {diclk_dfnblist}")

        # --fclk
        high_nb_len = len(high_nb)
        low_nb_len = len(low_nb)
        df_fcknb125low = df.loc[
            (clk_src)
            & (df["pvt"].str[0:high_nb_len] == high_nb)
            & (df["pulse"] == "(low)")
        ]
        df_fcknbn40low = df.loc[
            (clk_src)
            & (df["pvt"].str[0:low_nb_len] == low_nb)
            & (df["pulse"] == "(low)")
        ]
        df_fcknb125hi = df.loc[
            (clk_src)
            & (df["pvt"].str[0:high_nb_len] == high_nb)
            & (df["pulse"] == "(high)")
        ]
        df_fcknbn40hi = df.loc[
            (clk_src)
            & (df["pvt"].str[0:low_nb_len] == low_nb)
            & (df["pulse"] == "(high)")
        ]

        df_fcknb125lowst = df_fcknb125low[
            df_fcknb125low.tpwda == df_fcknb125low.tpwda.min()
        ].head(1)
        df_fcknbn40lowst = df_fcknbn40low[
            df_fcknbn40low.tpwda == df_fcknbn40low.tpwda.min()
        ].head(1)
        df_fcknb125hist = df_fcknb125hi[
            df_fcknb125hi.tpwda == df_fcknb125hi.tpwda.min()
        ].head(1)
        df_fcknbn40hist = df_fcknbn40hi[
            df_fcknbn40hi.tpwda == df_fcknbn40hi.tpwda.min()
        ].head(1)

        fclk_dfnblist = [
            df_fcknb125lowst,
            df_fcknbn40lowst,
            df_fcknb125hist,
            df_fcknbn40hist,
        ]
        dprint(HIGH, f"fclk_dfnblist: {fclk_dfnblist}")

    # Boost
    if high_b != "NA" and low_b != "NA":

        # --diclk
        high_b_len = len(high_b)
        low_b_len = len(low_b)
        df_dckb125low = df.loc[
            (clk_src)
            & (df["pvt"].str[0:high_b_len] == high_b)
            & (df["pulse"] == "(low)")
        ]
        df_dckbn40low = df.loc[
            (clk_src)
            & (df["pvt"].str[0:low_b_len] == low_b)
            & (df["pulse"] == "(low)")
        ]
        df_dckb125hi = df.loc[
            (clk_src)
            & (df["pvt"].str[0:high_b_len] == high_b)
            & (df["pulse"] == "(high)")
        ]
        df_dckbn40hi = df.loc[
            (clk_src)
            & (df["pvt"].str[0:low_b_len] == low_b)
            & (df["pulse"] == "(high)")
        ]

        df_dckb125lowst = df_dckb125low[
            df_dckb125low.tpwda == df_dckb125low.tpwda.min()
        ].head(1)
        df_dckbn40lowst = df_dckbn40low[
            df_dckbn40low.tpwda == df_dckbn40low.tpwda.min()
        ].head(1)
        df_dckb125hist = df_dckb125hi[
            df_dckb125hi.tpwda == df_dckb125hi.tpwda.min()
        ].head(1)
        df_dckbn40hist = df_dckbn40hi[
            df_dckbn40hi.tpwda == df_dckbn40hi.tpwda.min()
        ].head(1)

        diclk_dfblist = [
            df_dckb125lowst,
            df_dckbn40lowst,
            df_dckb125hist,
            df_dckbn40hist,
        ]

        # --fclk
        high_b_len = len(high_b)
        low_b_len = len(low_b)
        df_fckb125low = df.loc[
            (clk_src)
            & (df["pvt"].str[0:high_b_len] == high_b)
            & (df["pulse"] == "(low)")
        ]
        df_fckbn40low = df.loc[
            (clk_src)
            & (df["pvt"].str[0:low_b_len] == low_b)
            & (df["pulse"] == "(low)")
        ]
        df_fckb125hi = df.loc[
            (clk_src)
            & (df["pvt"].str[0:high_b_len] == high_b)
            & (df["pulse"] == "(high)")
        ]
        df_fckbn40hi = df.loc[
            (clk_src)
            & (df["pvt"].str[0:low_b_len] == low_b)
            & (df["pulse"] == "(high)")
        ]
        df_fckb125lowst = df_fckb125low[
            df_fckb125low.tpwda == df_fckb125low.tpwda.min()
        ].head(1)
        df_fckbn40lowst = df_fckbn40low[
            df_fckbn40low.tpwda == df_fckbn40low.tpwda.min()
        ].head(1)
        df_fckb125hist = df_fckb125hi[
            df_fckb125hi.tpwda == df_fckb125hi.tpwda.min()
        ].head(1)
        df_fckbn40hist = df_fckbn40hi[
            df_fckbn40hi.tpwda == df_fckbn40hi.tpwda.min()
        ].head(1)

        fclk_dfblist = [
            df_fckb125lowst,
            df_fckbn40lowst,
            df_fckb125hist,
            df_fckbn40hist,
        ]

    wb = load_workbook(DCDoutput_wb)
    ws = wb[DCDoutput_ws]

    ws_names = []
    ws_tables = []

    for names in ws.tables:
        ws_names.append(names)

    ws_tables.append(ws.tables[ws_names[0]])
    ws_tables.append(ws.tables[ws_names[1]])
    diclkTable = ws_tables[0]
    fclkTable = ws_tables[1]

    # diclk----------------------
    diclkRange = diclkTable.ref

    dprint(MEDIUM, f"DICLK: {diclkRange}")

    diclkStartEnd = diclkRange.split(":")

    diclkStartCell = diclkStartEnd[0]
    diclkEndCell = diclkStartEnd[1]

    temp = re.compile("([a-zA-Z]+)([0-9]+)")
    diresStart = temp.match(diclkStartCell).groups()
    diresEnd = temp.match(diclkEndCell).groups()

    dicolumnStart = diresStart[0]
    dirowEnd = int(diresEnd[1])

    fclkRange = fclkTable.ref

    dprint(MEDIUM, f"FCLK: {fclkRange}")
    fclkStartEnd = fclkRange.split(":")

    fclkStartCell = fclkStartEnd[0]
    fclkEndCell = fclkStartEnd[1]

    fresStart = temp.match(fclkStartCell).groups()
    fresEnd = temp.match(fclkEndCell).groups()

    fcolumnStart = fresStart[0]
    frowEnd = int(fresEnd[1])

    # -------------------------------------------------------------------------
    # ----------------------------------diclk----------------------------------
    # -------------------------------------------------------------------------

    # Boost

    currentRowEnd = 0

    clk_status = "di"
    boost_status = "b"
    if high_b != "NA" and low_b != "NA":
        nprint("ppmin for diclk boost:")
        writer_f(
            diclk_dfblist,
            mc_diclk_high_boost,
            mc_diclk_low_boost,
            di_margin,
            dicolumnStart,
            dirowEnd,
            currentRowEnd,
            projectName,
            ws,
            boost_status,
            clk_status,
        )
        currentRowEnd = currentRowEnd + 4

    # Non Boost
    clk_status = "di"
    boost_status = "nb"
    if high_nb != "NA" and low_nb != "NA":
        nprint("ppmin for diclk non-boost:")
        writer_f(
            diclk_dfnblist,
            mc_diclk_high_nonboost,
            mc_diclk_low_nonboost,
            di_margin,
            dicolumnStart,
            dirowEnd,
            currentRowEnd,
            projectName,
            ws,
            boost_status,
            clk_status,
        )
        currentRowEnd = currentRowEnd + 4

    # -------------------------------------------------------------------------
    # ----------------------------------fclk-----------------------------------
    # -------------------------------------------------------------------------

    # Boost

    fcurrentRowEnd = 0
    clk_status = "f"
    boost_status = "b"
    if high_b != "NA" and low_b != "NA":
        nprint("ppmin for fclk boost")
        writer_f(
            fclk_dfblist,
            mc_fclk_high_boost,
            mc_fclk_low_boost,
            di_margin,
            fcolumnStart,
            frowEnd,
            fcurrentRowEnd,
            projectName,
            ws,
            boost_status,
            clk_status,
        )
        fcurrentRowEnd = fcurrentRowEnd + 4

    # Non Boost
    boost_status = "nb"
    clk_status = "f"
    if high_nb != "NA" and low_nb != "NA":
        print("ppmin for fclk non-boost:")
        writer_f(
            fclk_dfnblist,
            mc_fclk_high_nonboost,
            mc_fclk_low_nonboost,
            di_margin,
            fcolumnStart,
            frowEnd,
            fcurrentRowEnd,
            projectName,
            ws,
            boost_status,
            clk_status,
        )
        fcurrentRowEnd = fcurrentRowEnd + 4

    diclkTable.ref = (
        diclkStartCell + ":" + diresEnd[0] + str(dirowEnd + currentRowEnd)
    )
    fclkTable.ref = (
        fclkStartCell + ":" + fresEnd[0] + str(frowEnd + fcurrentRowEnd)
    )

    wb.save(DCDoutput_wb)
    # -------------------------------------------------------------------------


if __name__ == "__main__":

    args = parse_args()
    filename = os.getcwd() + "/" + os.path.basename(__file__) + ".log"
    logger = create_logger(filename)  # Create log file

    # Register exit function
    atexit.register(footer)

    # Initalise shared variables and run main
    version = get_release_version()
    CommonHeader.init(args, __author__, version)

    header()
    main(args)
    iprint(f"Log file: {filename}")
    utils__script_usage_statistics(__tool_name__, version)
