#!/depot/Python/Python-3.8.0/bin/python
"""
Name    : ocv.py
Author  : Angelina Chan
Date    : Jan 16, 2023
Purpose : Find the OCV from Monte Carlo simulation data and append
          to the end of the tables in Aging_DCD.../(OutputSheet) for:

Parameters:

1. PCLk and TxClk OR DICLk and FClk (rxclk_outdi & rxclk_dqrxflop)
2. High(Temp) and Low(Temp)
3  orient (macro orientation)
3. Boost and/or Non-Boost
4. MC and DI Margin

Instructions on usage:

1. Download MPW & Aging_DCD xlsx sheets in current path
2. Edit the parameters in the config file (multiple projects supported)
3. Run using ocv.py -c config_aging.ini
4. This script will append the data to the end of the table defined in the
   Aging_DCD excel file
NOTE: PCLK or RX must be in the config section names (e.g, [D910 - PCLK] ).
      Can use the same config file/parameters as Aging_DCD_output_*.py
"""

import argparse
import atexit
import os
import pathlib
import sys
from typing import List

import pandas as pd
import configparser
import math
import re
from openpyxl import load_workbook


# ---------------------------------- #
bindir = str(pathlib.Path(__file__).resolve().parent)
# Add path to library that may be symbolically linked.
sys.path.append(bindir + "/../lib/Util")
# Add path to sharedlib's Python Utilities directory.
sys.path.append(bindir + "/../lib/python/Util")
# Add path to Aging package directory.
sys.path.append(bindir + "/../lib/python")
# ---------------------------------- #

from CommonHeader import LOW, MEDIUM, HIGH
from Messaging import iprint, fatal_error, nprint
from Messaging import viprint, dprint
from Messaging import create_logger, footer, header

from Misc import utils__script_usage_statistics, get_release_version
import CommonHeader

import aging_pkg

__author__ = "Angelina Chan"
__tool_name__ = "ocv"


def parse_args():

    # Always include -v, and -d
    parser = argparse.ArgumentParser(
        description=(
            "Find the OCV from Monte Carlo simulation data and"
            " append to the end of the tables in Aging_DCD.../"
            "(OutputSheet) for:\n\nParameters:\n1. PCLk and TxClk"
            " OR DICLk and FClk (rxclk_outdi & rxclk_dqrxflop) "
            "\n2. High(Temp) and Low(Temp)\n3. "
            "Boost and/or Non-Boost\n4. MC Margin\n"
            "Leave unused parameters in config as 'NA' or empty\n\n"
            "Instructions on usage:\n\n1. Download "
            "Aging_DCD xlsx sheets in current path\n2. "
            "Edit the parameters in the config file "
            "(multiple projects supported)\n3. Run using ..python "
            "ocv.py -c config_aging.ini\n4. "
            "This script will append the data to the end of the "
            "table defined in the Aging_DCD excel file.\n"
            "NOTE: PCLK or RX must be in the config section names "
            "(e.g, [D910 - PCLK] ). Can use the same config file/"
            "parameters as Aging_DCD_output_*.py"
        ),
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "-v", metavar="<#>", type=int, default=0, help="verbosity"
    )
    parser.add_argument("-d", metavar="<#>", type=int, default=0, help="debug")

    # -------------------------------------
    parser.add_argument(
        "-c", "--cfile", help="REQUIRED. Config file path", required=True
    )
    # -------------------------------------
    args = parser.parse_args()
    return args


def main(args):
    aging_pkg.check_config_exists(args.cfile)
    project_list = parse_config(args.cfile)
    for proj in project_list:
        if not os.path.isfile(proj.DCDoutput_wb):
            fatal_error(f"File not found, '{proj.DCDoutput_wb}'. Exiting...")
        iprint(f"MC Data used for project: {proj.projectName} - {proj.clock}")

        # --Boost
        if proj.mc_data_b != "NA":
            df_b = pd.read_excel(proj.mc_data_b, engine="openpyxl")
            mc_clk1_lo_b = mc_final(df_b, proj.temp_low, proj.matches_1, proj)
            nprint(f"\tmc_{proj.clock_1}_low_boost: {mc_clk1_lo_b}")
            mc_clk1_hi_b = mc_final(df_b, proj.temp_high, proj.matches_1, proj)
            nprint(f"\tmc_{proj.clock_1}_high_boost: {mc_clk1_hi_b}")
            mc_clk2_lo_b = mc_final(df_b, proj.temp_low, proj.matches_2, proj)
            nprint(f"\tmc_{proj.clock_2}_low_boost: {mc_clk2_lo_b}")
            mc_clk2_hi_b = mc_final(df_b, proj.temp_high, proj.matches_2, proj)
            nprint(f"\tmc_{proj.clock_2}_high_boost: {mc_clk2_hi_b}")
            freq_b = get_freq(df_b)
        else:
            mc_clk1_lo_b = 0
            mc_clk1_hi_b = 0
            mc_clk2_lo_b = 0
            mc_clk2_hi_b = 0
            freq_b = 0

        # --Non Boost
        if proj.mc_data_nb != "NA":
            df_nb = pd.read_excel(proj.mc_data_nb, engine="openpyxl")
            mc_clk1_lo_nb = mc_final(df_nb, proj.temp_low, proj.matches_1, proj)
            nprint(f"\tmc_{proj.clock_1}_low_nonboost: {mc_clk1_lo_nb}")
            mc_clk1_hi_nb = mc_final(
                df_nb, proj.temp_high, proj.matches_1, proj
            )
            nprint(f"\tmc_{proj.clock_1}_high_nonboost: {mc_clk1_hi_nb}")
            mc_clk2_lo_nb = mc_final(df_nb, proj.temp_low, proj.matches_2, proj)
            nprint(f"\tmc_{proj.clock_2}_low_nonboost: {mc_clk2_lo_nb}")
            mc_clk2_hi_nb = mc_final(
                df_nb, proj.temp_high, proj.matches_2, proj
            )
            nprint(f"\tmc_{proj.clock_2}_high_nonboost: {mc_clk2_hi_nb}")
            freq_nb = get_freq(df_nb)
        else:
            mc_clk1_lo_nb = 0
            mc_clk1_hi_nb = 0
            mc_clk2_lo_nb = 0
            mc_clk2_hi_nb = 0
            freq_nb = 0

        # #####################################################################
        workbook(
            proj,
            freq_nb,
            freq_b,
            mc_clk2_hi_b,
            mc_clk2_lo_b,
            mc_clk2_hi_nb,
            mc_clk2_lo_nb,
            mc_clk1_hi_b,
            mc_clk1_lo_b,
            mc_clk1_hi_nb,
            mc_clk1_lo_nb,
        )


class agingData:
    def __init__(
        self,
        clk: str,
        DCDout_wb: str,
        DCDout_ws: str,
        projName: str,
        param_percent: float,
        mc_data_b: str,
        mc_data_nb: str,
        temp_lo: int,
        temp_hi: int,
        automotive: bool,
    ) -> None:
        self.clock = clk
        self.DCDoutput_wb = DCDout_wb
        self.DCDoutput_ws = DCDout_ws
        self.projectName = projName
        self.param_percentage = param_percent

        self.mc_data_b = mc_data_b
        self.mc_data_nb = mc_data_nb
        self.temp_low = temp_lo
        self.temp_high = temp_hi

        self.automotive = automotive
        # Only use the monte results from negative pulse
        if "pclk" in clk:
            self.matches_1 = ["txclk_dqs_pn", "std"]
            self.matches_2 = ["lcdl_in_pn", "std"]
            self.clock_1 = "Txclk"
            self.clock_2 = "pclk"
            self.temp = "tfresh"
        elif "rx" in clk:
            self.matches_1 = ["rxclk_dqrxflop_pn", "std"]
            self.matches_2 = ["rxclk_outdi_pn", "std"]
            self.clock_1 = "fclk"
            self.clock_2 = "diclk"
            self.temp = "temp"
        else:
            fatal_error(f"Invalid clock: {clk}")


def parse_config(config_file) -> List[agingData]:
    """Parses the config file"""
    """
    ======================================================================
    Sample config:
    ======================================================================
    Instructions:
    1. The default section sets the default values to NA unless redefined under a new section.
    2. Can add as many project sections to run.

    [DEFAULT]
    DCDoutput_wb = NA  # Output excel path
    DCDoutput_ws = NA  # Output sheet name
    orient = NA  # Orientation
    projectName = NA  # project name
    di_margin = NA  # DISCLACK margin
    param_percentage = 2.5
    high_nb = NA  # Non-Boost corner high temp
    low_nb = NA  # Non-Boost corner low temp
    high_b = NA  # Boost corner high temp
    low_b = NA  # Boost corner low temp
    mc_data_nb = NA  # Monte-carlo non-boost data path
    mc_data_b = NA  # Monte-carlo boost data path
    temp_low = -40  # Low temp
    temp_high = 125   # High temp
    automotive_project = False  # Is automotive project?

    [D932 0p65 - PCLK]
    DCDoutput_wb = d932DISLACK.xlsx
    DCDoutput_ws = d932DI
    orient = ew
    high_nb = ssgnp0p585v125
    low_nb = ssgnp0p585vn40
    high_b = NA
    low_b = NA
    projectName = D9320p65
    di_margin = 2
    mc_data_nb = tb_pclk_path_monte_0p65_1867.xlsx
    mc_data_b = tb_pclk_path_monte_0p65_1867.xlsx
    temp_low = -40
    temp_high = 125
    automotive_project = False
    """
    config = configparser.ConfigParser()
    config.read(config_file)

    data_list = []
    for project in config.sections():
        DCDoutput_wb = config[project]["DCDoutput_wb"]
        DCDoutput_ws = config[project]["DCDoutput_ws"]
        projectName = config[project]["projectName"]
        param_percentage = float(config[project]["param_percentage"])

        mc_data_b = config[project]["mc_data_b"]
        mc_data_nb = config[project]["mc_data_nb"]
        temp_low = config[project]["temp_low"]
        temp_high = config[project]["temp_high"]
        automotive_project = config[project]["automotive_project"]

        automotive_proj = False
        if automotive_project.lower() in ["true", "1", "t", "y", "yes"]:
            automotive_proj = True
        if "pclk" in project.lower():
            clock = "pclk"
        elif "rx" in project.lower():
            clock = "rx"
        else:
            fatal_error("No RX or PCLK in config section.")
        # #####################################################################
        if mc_data_b == "NA" and mc_data_nb == "NA":
            fatal_error(
                (
                    "ERROR:\tBoth Monte Carlo boost and non-boost are not "
                    "defined - Check config. Exiting..."
                )
            )
        data = agingData(
            clock,
            DCDoutput_wb,
            DCDoutput_ws,
            projectName,
            param_percentage,
            mc_data_b,
            mc_data_nb,
            temp_low,
            temp_high,
            automotive_proj,
        )
        data_list.append(data)
    return data_list


# ########################## FUNCTIONS FOR MC OUTPUT ##########################
def mc_out(array, temp, df, proj: agingData):

    dataFrameArr = ["mos", proj.temp] + array
    clk_df = df.filter(dataFrameArr, axis=1)
    dprint(LOW, f"TEMP:\n{temp}")
    # clk_df_x = clk_df[(clk_df['mos'] == 'mos_ss')
    # & (clk_df['tfresh'] == temp)]
    dprint(HIGH, f"CLK_DF:\n{clk_df}")
    clk_df[proj.temp] = clk_df[proj.temp].astype(str)
    clk_df_x = clk_df[(clk_df[proj.temp] == str(temp))]
    dprint(HIGH, f"CLK_DF_X:\n{clk_df_x}")
    try:
        clk_row_x = clk_df_x.iloc[0]
    except IndexError as error:
        fatal_error(
            (
                f"No matching temp for {temp} in data. Please check config"
                f" file or data for errors. {repr(error)}"
            )
        )
    clk_row_x[array] = clk_row_x[array].apply(pd.to_numeric)

    max_fromdf = pd.DataFrame(clk_row_x[array])
    mc_number = float(max_fromdf.max(axis=0).iloc[0])

    if proj.automotive:
        # monte value used for automotive project changed from std*5 to std*3
        returnvalue = round(mc_number * 3, 2)

    else:
        returnvalue = round(mc_number * 3, 2)

    return returnvalue


def mc_final(df, temp, matches, proj):

    array_clk = aging_pkg.columnArrOut(df, matches)
    mc_clk = mc_out(array_clk, temp, df, proj)

    return mc_clk


# #############################################################################


def get_freq(df):
    # Frequency in MHz
    if "freq" in df:
        var = "freq"
    elif "clkfreq" in df:
        var = "clkfreq"
    elif "bitrate" in df:
        # freq = bitrate/2
        var = "bitrate"
    dprint(LOW, f"Getting frequency from {var} column")
    freq_df = df.filter([var], axis=1).dropna(axis=0)
    dprint(HIGH, f"\n{freq_df}")
    freq_element = float(freq_df[var][0])
    if var == "bitrate":
        freq_element = float(freq_element) / 2
    if len(str(freq_df[var][0]).split("e")) > 1:
        freq = freq_element / 1000000
    else:
        freq = freq_element
    dprint(LOW, f"\tFrequency: {freq}")
    return freq


def get_mpw(freq, ocv):
    period = 1 / freq
    mpw = period - ocv
    return mpw


def writer_f(
    freq, mc_high, mc_low, col_f, row_f, currentrow_f, proj: agingData, ws
):
    for i, j in zip(range(1, 5), range(1, 5)):

        projectnameLow = proj.projectName + " Low"
        projectnameHigh = proj.projectName + " High"

        if j > 2:
            projectStart = projectnameHigh
            # mc_f = mc_high
        else:
            projectStart = projectnameLow
            # mc_f = mc_low

        if (j % 2) == 0:
            mc_f = mc_low
            temp = str(proj.temp_low) + "C"
        else:
            mc_f = mc_high
            temp = str(proj.temp_high) + "C"

        period = 1000000 / freq  # in ps
        # param_percentage = 2.5 by default
        ix = proj.param_percentage * (period / 100)
        iy = math.sqrt(pow(mc_f, 2) + pow(ix, 2))
        ir = float(ix + iy)
        # -----------------------------------------
        rp = round(period, 2)
        rir = round(ir, 2)
        mpw = (period / 2) - rir
        rmpw = round(mpw, 2)
        # Project
        ws[(col_f) + str(row_f + i + currentrow_f)] = projectStart
        viprint(LOW, f"Project: {projectStart}")
        # Period of operation, ps
        ws[chr(ord(col_f) + 1) + str(row_f + i + currentrow_f)] = rp
        viprint(LOW, f"Period of operation, ps: {rp}")
        # Tempurature, C
        ws[chr(ord(col_f) + 2) + str(row_f + i + currentrow_f)] = temp
        viprint(LOW, f"Tempurature, C: {proj.temp}")
        # OCV, ps
        ws[chr(ord(col_f) + 4) + str(row_f + i + currentrow_f)] = rir
        viprint(LOW, f"OCV, ps: {rir}")
        # Target Minimum Input Pulse Width, ps
        ws[chr(ord(col_f) + 6) + str(row_f + i + currentrow_f)] = rmpw
        viprint(LOW, f"Target Minimum Input Pulse Width, ps: {rmpw}")


def workbook(
    proj: agingData,
    freq_nb,
    freq_b,
    mc_clk2_hi_b,
    mc_clk2_lo_b,
    mc_clk2_hi_nb,
    mc_clk2_lo_nb,
    mc_clk1_hi_b,
    mc_clk1_lo_b,
    mc_clk1_hi_nb,
    mc_clk1_lo_nb,
):
    wb = load_workbook(proj.DCDoutput_wb)
    ws = wb[proj.DCDoutput_ws]

    ws_names = []
    ws_tables = []

    for names in ws.tables:
        ws_names.append(names)

    ws_tables.append(ws.tables[ws_names[0]])
    ws_tables.append(ws.tables[ws_names[1]])
    clk2Table = ws_tables[0]
    clk1Table = ws_tables[1]

    # CLK2 (PCLK/DICLK)----------------------
    clk2Range = clk2Table.ref

    dprint(MEDIUM, f"{proj.clock_2}: {clk2Range}")

    clk2StartEnd = clk2Range.split(":")

    clk2StartCell = clk2StartEnd[0]
    clk2EndCell = clk2StartEnd[1]

    temp = re.compile("([a-zA-Z]+)([0-9]+)")
    c2resStart = temp.match(clk2StartCell).groups()
    c2resEnd = temp.match(clk2EndCell).groups()

    c2colStart = c2resStart[0]
    c2rowEnd = int(c2resEnd[1])

    # CLK1 (TXCLK/FCLK)----------------------
    clk1Range = clk1Table.ref

    dprint(MEDIUM, f"{proj.clock_1}: {clk1Range}")
    clk1StartEnd = clk1Range.split(":")

    clk1StartCell = clk1StartEnd[0]
    clk1EndCell = clk1StartEnd[1]

    c1resStart = temp.match(clk1StartCell).groups()
    c1resEnd = temp.match(clk1EndCell).groups()

    c1colStart = c1resStart[0]
    c1rowEnd = int(c1resEnd[1])

    # -------------------------------------------------------------------------
    # -------------------------------PXCLK/DICLK-------------------------------
    # -------------------------------------------------------------------------

    # Boost
    currentRowEnd = 0
    if mc_clk2_hi_b != 0 and mc_clk2_lo_b != 0:
        viprint(LOW, f"{proj.clock_2} boost:")
        writer_f(
            freq_b,
            mc_clk2_hi_b,
            mc_clk2_lo_b,
            c2colStart,
            c2rowEnd,
            currentRowEnd,
            proj,
            ws,
        )
        currentRowEnd = currentRowEnd + 4

    # Non Boost
    if mc_clk2_hi_nb != 0 and mc_clk2_lo_nb != 0:
        viprint(LOW, f"{proj.clock_2} non-boost:")
        writer_f(
            freq_nb,
            mc_clk2_hi_nb,
            mc_clk2_lo_nb,
            c2colStart,
            c2rowEnd,
            currentRowEnd,
            proj,
            ws,
        )
        currentRowEnd = currentRowEnd + 4

    # -------------------------------------------------------------------------
    # ----------------------------------TXCLK----------------------------------
    # -------------------------------------------------------------------------

    # Boost
    tcurrentRowEnd = 0
    if mc_clk1_hi_b != 0 and mc_clk1_lo_b != 0:
        viprint(LOW, f"{proj.clock_1} boost:")
        writer_f(
            freq_b,
            mc_clk1_hi_b,
            mc_clk1_lo_b,
            c1colStart,
            c1rowEnd,
            tcurrentRowEnd,
            proj,
            ws,
        )
        tcurrentRowEnd = tcurrentRowEnd + 4

    # Non Boost
    if mc_clk1_hi_nb != 0 and mc_clk1_lo_nb != 0:
        viprint(LOW, f"{proj.clock_1} non-boost:")
        writer_f(
            freq_nb,
            mc_clk1_hi_nb,
            mc_clk1_lo_nb,
            c1colStart,
            c1rowEnd,
            tcurrentRowEnd,
            proj,
            ws,
        )
        tcurrentRowEnd = tcurrentRowEnd + 4

    clk2Table.ref = (
        clk2StartCell + ":" + c2resEnd[0] + str(c2rowEnd + currentRowEnd)
    )
    clk1Table.ref = (
        clk1StartCell + ":" + c1resEnd[0] + str(c1rowEnd + tcurrentRowEnd)
    )

    wb.save(proj.DCDoutput_wb)
    # -------------------------------------------------------------------------


if __name__ == "__main__":

    args = parse_args()
    filename = os.getcwd() + "/" + os.path.basename(__file__) + ".log"
    logger = create_logger(filename)  # Create log file

    # Register exit function
    atexit.register(footer)

    # Initalise shared variables and run main
    version = get_release_version()
    CommonHeader.init(args, __author__, version)

    header()
    main(args)
    iprint(f"Log file: {filename}")
    utils__script_usage_statistics(__tool_name__, version)
