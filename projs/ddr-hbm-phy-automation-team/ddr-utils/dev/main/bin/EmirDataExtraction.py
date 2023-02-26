#!/depot/Python/Python-3.8.0/bin/python
# nolint main
import pandas as pd
import configparser
from openpyxl import load_workbook
import os
from tkinter import ttk
import tkinter as tk
import getpass
import sys
from itertools import chain
import subprocess
from pathlib import Path

# GUI

BIN_DIR = str(Path(__file__).resolve().parent)
# Add path to sharedlib's Python Utilities directory.
sys.path.append(BIN_DIR + "/../lib/python/Util")

import Misc


Misc.utils__script_usage_statistics("EmirDataExtraction.py", "2022ww24")

config = configparser.RawConfigParser()


def resize(event):
    canvas.coords(line, 0, 0, event.width, event.height)
    canvas.bind("<Configure>", resize)


root = tk.Tk()
container = ttk.Frame(root)

container.pack(fill=tk.BOTH, expand=tk.YES)

canvas = tk.Canvas(container)

canvas.pack(side="left", fill="both", expand=True)
line = canvas.create_line(0, 0, 0, 0)

scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
scrollable_frame = ttk.Frame(canvas)
scrollbar.pack(side="right", fill="y")

scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

canvas.configure(yscrollcommand=scrollbar.set)


v = tk.IntVar()
v.set(0)
x = tk.IntVar()
x.set(0)
m = tk.IntVar()
m.set(0)
root.wm_title("EMIR Extraction")
cwd = os.getcwd()
path = tk.StringVar()
macro_s = tk.StringVar()
layer_s = tk.StringVar()
password = tk.StringVar()

# shows the input choices for the em and ir


def ShowChoice():
    # print('v is :',v.get(),'x is :',x.get())
    # print('m is :',m.get())
    return (v.get(), x.get())


def combine_funcs(*funcs):
    def combined_func(*args, **kwargs):
        for f in funcs:
            f(*args, **kwargs)

    return combined_func


def delete_entry():
    global myEntry2
    myEntry.delete(0, tk.END)
    myEntry2.delete(0, tk.END)


pathcounter = 0


def get_entry():
    global pathcounter
    print("Entered path is:", path.get())
    get_entry.string = path.get()
    pathcounter = pathcounter + 1
    enterEntry.destroy()


def check_pathcounter():
    if pathcounter == 1:
        pwds()


def get_macro():
    global enterEntry_macro
    print("Entered macro names:", macro_s.get())
    get_macro.string = macro_s.get()
    enterEntry_macro.destroy()


def get_layer():
    print("Entered metal layer name:", layer_s.get())
    get_layer.string = layer_s.get()


pwcounter = 0


def get_pw():
    global pwcounter
    if pwcounter == 1:
        get_entry()
    print("Enter password for user:", password.get().replace(password.get(), "*"))
    get_pw.string = password.get()
    pwcounter = pwcounter + 1


def check_pwcounter():
    global enterEntry2
    if pwcounter == 2:
        enterEntry2.destroy()


tblist_gui = []
macrolist_gui = []
modelist_gui = []


download_l = []
username_l = []
pwd_l = []
macro_list_l = []
layer_select_l = []
cleaned_path_l = []

using_p4 = [False]
p4_root = []


def find_tb():
    cwd = os.getcwd()
    # global download,username,pwd, macro_list,layer_select
    if len(sys.argv) > 1:
        layer_select = process_args()
    else:
        cleaned_path = get_entry.string.split(".xlsx")[
            0
        ]  # it cutes everything after .xlsx including itself
        macro_string = get_macro.string
        macro_string = macro_string.split("\n")[0]
        macro_list = [x.strip() for x in macro_string.split(",")]
        username = getpass.getuser()
        pwd = get_pw.string
        cleaned_path = cleaned_path + ".xlsx"
        layer_select = get_layer.string

    username_l.append(username)
    pwd_l.append(pwd)
    cleaned_path_l.append(cleaned_path)
    macro_list_l.append(macro_list)
    layer_select_l.append(layer_select)

    file_path = cleaned_path
    file_array = file_path.split("/")
    file_name = file_array[-1]

    test = os.listdir(cwd)
    for item in test:
        if item.startswith(file_name):
            #   #print('deleted ',file_name)
            os.remove(
                os.path.join(cwd, item)
            )  # delete the old xlsx file generated to prevent overwrite

    # if from perforce
    if cleaned_path.startswith("//"):

        using_p4[0] = True

        print("downloading from perforce")
        p4_root.append(cleaned_path)
        command = "p4 sync -f " + cleaned_path
        locate_p4 = "p4 where " + cleaned_path
        subprocess.check_output(command, shell=True)
        p4_path = subprocess.check_output(locate_p4, shell=True, universal_newlines=True)

        xlsx_path = str(p4_path.split()[2])
        os.system("chmod 777 " + xlsx_path)
        p4_root.append(xlsx_path)
        os.system("cp " + xlsx_path + " .")
    else:
        print("downloading from sharepoint")
        download = (
            "wget -q --user=" + username + " --password=" + "'" + pwd + "'" + " " + cleaned_path
        )  # downloading the excel file from the sharepoint
        print(download)
        os.system(download)

        download_l.append(download)

    emir_sheet = pd.read_excel(cwd + "/" + file_name, "EMIR", header=None)
    emir_sheet.iloc[:, 0] = emir_sheet.iloc[:, 0].fillna(
        method="ffill"
    )  # it assigns the name of the merged macro name cells to all of the cells withins it
    load_workbook(filename=file_name)
    bottom = emir_sheet.iloc[:, 0].tail(
        1
    )  # gets the last element in the first column of excel file,will print 62
    bottom = list(bottom.index.values)  # will print [62]
    last_row = bottom[0]  # will print 62
    a = 3

    while a <= last_row:
        macro_name = emir_sheet.iloc[a][0]
        mode_name = emir_sheet.iloc[a][4]
        tb_name = emir_sheet.iloc[a][5]
        if macro_name in macro_list:

            if not isinstance(tb_name, float):
                dirpath = cwd + "/" + macro_name + "/" + tb_name

                if os.path.exists(dirpath):
                    macrolist_gui.append(macro_name)
                    tblist_gui.append(tb_name)
                    modelist_gui.append(mode_name)

        a = a + 1


def process_args():
    """Called if len(sys.argv) > 1"""
    config_name = sys.argv[2]

    with open(config_name, "r+") as f_config:
        config_string = "[EMIR]\n" + f_config.read()
        config.read_string(config_string)
        f_config.close()
    L = config.sections()

    for element in L:
        try:
            cleaned_path = config.get(element, "sharepoint_path")
            cleaned_path = cleaned_path.split(".xlsx")[0]
            cleaned_path = cleaned_path + ".xlsx"

            macro_list = config.get(element, "macro_name")
            macro_list = [x.strip() for x in macro_list.split(",")]

            layer_select = config.get(element, "metal_layer")

        except Exception as e:
            print("ERROR:The config file has missing inputs and/or the format is incorrect ")
            print(e)
            help()
            sys.exit()

        username = getpass.getuser()
        print("Enter password for", username, ":")
        pwd = getpass.getpass("")

        f_download = open("errorcheck.txt", "w+")
        checkconfig = (
            "curl --ntlm -k --user "
            + username
            + ":"
            + "'"
            + pwd
            + "'"
            + " --head --fail "
            + cleaned_path
            + ' --output "errorcheck.txt" '
        )  # checks the correctness of pw
        os.system(checkconfig)

        f_download.close()

    with open("errorcheck.txt", "r") as f:
        valid = False
        for line in f.readlines():
            if "200 OK" in line:
                valid = True
        if not valid:
            print(
                "ERROR:Invalid password or/and url, please check the password and the URL inside the .ini file again."
            )
            print("Please try again")
            sys.exit()
    return layer_select


func_counter = 0


def checkandprintmacros():  # noqa: C901
    cwd = os.getcwd()
    global func_counter

    if func_counter == 1 and len(sys.argv) > 1:
        root.destroy()
        # root.mainloop()
    if len(sys.argv) > 1 and func_counter == 0:
        find_tb()
        func_counter += 1

        emir_summary()
        sys.exit()

    username = ", ".join(username_l)
    pwd = ", ".join(pwd_l)
    download = ", ".join(download_l)
    layer_select = ", ".join(layer_select_l)
    macro_list = list(chain.from_iterable(macro_list_l))

    print(using_p4)

    if not using_p4[0]:
        cleaned_path = download.split()[-1]
        file_path = cleaned_path
    else:
        file_path = p4_root[0]

    file_array = file_path.split("/")
    file_name = file_array[-1]

    if len(sys.argv) == 1 or (len(sys.argv) > 1 and func_counter == 1):

        EMcorner_list = [i.split(",") for i in EMcornerlist]

        for i in range(len(EMcorner_list)):
            print(
                'Entered EM exlcude corners for macro "'
                + macrolist_gui[i]
                + '", testbench "'
                + tblist_gui[i]
                + '", "'
                + str(modelist_gui[i])
                + '" are: '
                + EMcornerlist[i]
            )
            print(
                'Entered IR exlcude corners for macro "'
                + macrolist_gui[i]
                + '", testbench "'
                + tblist_gui[i]
                + '", "'
                + str(modelist_gui[i])
                + '" are: '
                + IRcornerlist[i]
            )

        tb_corner_ex = list(map(list, zip(tblist_gui, modelist_gui, EMcornerlist, IRcornerlist)))

        tb_corner_ex = [i for i in tb_corner_ex if "all" not in i]

        tb_corner_ex = list(zip(*tb_corner_ex))

        tb_list = tb_corner_ex[0]
        tb_list = [i for i in tb_list]

        mode_list = tb_corner_ex[1]
        mode_list = [i for i in mode_list]

        EM_per_tb = tb_corner_ex[2]
        EM_per_tb = [i for i in EM_per_tb]  # convert tuple to list
        EMlist_per_tb = [
            i.split(",") for i in EM_per_tb
        ]  # covert to list of list  # [['2', '4'], ['6'], ['2'], ['3'], ['5']] this is used for diff mode with same tb

        IR_per_tb = tb_corner_ex[3]
        IR_per_tb = [i for i in IR_per_tb]  # convert tuple to list
        IRlist_per_tb = [i.split(",") for i in IR_per_tb]  # covert to list of list

    # print(tb_list)
    # print(mode_list)
    # print(EMlist_per_tb) # [['2', '4'], ['6'], ['2'], ['3'], ['5']]
    # print(IRlist_per_tb)

    if len(sys.argv) == 1:
        root.destroy()

    # THE PATH FOR UPLOADING FILE
    emir_sheet = pd.read_excel(cwd + "/" + file_name, "EMIR", header=None)
    emir_sheet.iloc[:, 0] = emir_sheet.iloc[:, 0].fillna(
        method="ffill"
    )  # it assigns the name of the merged macro name cells to all of the cells withins it
    wb = load_workbook(filename=file_name)
    sheet = wb["EMIR"]

    second_row = emir_sheet.iloc[1, :]  # SECOND ROW OF EXCEL TEMPLATE
    second_row = list(second_row)
    measurements_headers = emir_sheet.iloc[0, :]
    measurements_headers = list(measurements_headers)
    measurements = []

    EM_param = []
    IR_param = []
    for measurement in [x for x in second_row if not isinstance(x, float)]:
        measurement = measurement.split()
        measurement = "".join(measurement)
        if any(x for x in ["Iavg", "Irms", "acpc", "_wdt", "tmideg"] if x in measurement):
            measurements.append(measurement)  # you push your element here,
            EM_param.append(measurement)
        if "mV" in measurement:
            measurements.append(measurement)  # you push your element here,
            IR_param.append(measurement)
        if "Clean" in measurement:
            measurements.append(measurement)

    # print(measurements)

    column_counter = 0
    # gets the last element in the first column of excel file,will print 62
    bottom = emir_sheet.iloc[:, 0].tail(1)
    bottom = list(bottom.index.values)  # will print [62]
    last_row = bottom[0]  # will print 62
    i = 3
    check_list = []

    tb_counter = 0

    os.chdir(cwd)
    with open("erroroutput.log", "w+") as errormsgs:
        while i <= last_row:
            column_counter = 6
            macro_name = emir_sheet.iloc[i][0]
            tb_name = emir_sheet.iloc[i][5]
            mode_name = emir_sheet.iloc[i][4]
            if macro_name in macro_list:
                check_list.append(macro_name)

                if tb_name in tb_list:

                    if not isinstance(tb_name, float):
                        if ".bbSim" in tb_name:
                            tb_name = tb_name[:-6]
                        else:
                            tb_name == tb_name

                        dirpath = cwd + "/" + macro_name + "/" + tb_name

                        if os.path.exists(dirpath):

                            if mode_list[tb_counter] == mode_name:

                                print(
                                    "............................................................................................"
                                )
                                print("Macro name :", macro_name, "\t", "Testbench name:", tb_name)

                                for header in measurements_headers:
                                    if (
                                        not isinstance(header, float) and "Reports Path" in header
                                    ):  # writes to report path in excel file
                                        sheet.cell(
                                            row=i + 1,
                                            column=measurements_headers.index(header) + 1,
                                            value=dirpath,
                                        )

                                os.chdir(dirpath)  # changes the directory to dirpath

                                EM_onlist = []
                                IR_onlist = []

                                for file in os.listdir(dirpath):
                                    if "gz" in file and "ascii" in file:
                                        os.system("gunzip -f " + file)

                                os.system(
                                    "/remote/proj/alpha/alpha_common/bin/alphaXaSummary *ascii*"
                                )

                                os.system(
                                    "/remote/cad-rep/projects/alpha/alpha_common/bin/alphaXaWdtSummary.py"
                                )

                                existed_element = []
                                layerlist = []

                                EM_this_tb = EMlist_per_tb[tb_counter]
                                IR_this_tb = IRlist_per_tb[tb_counter]

                                if tb_counter == (
                                    len(EMlist_per_tb) - 1
                                ):  # increment the tb_counter if not the end of this em tb list

                                    tb_counter == tb_counter
                                else:
                                    tb_counter = tb_counter + 1

                                # print(tb_counter)
                                # print(mode_name)
                                for element in measurements:

                                    if not isinstance(element, float):
                                        element = (
                                            element.lower()
                                        )  # the files names are only in lower case
                                        element = element.split()
                                        element = "".join(element)

                                    em_dtod = []
                                    em_dtmos = []
                                    # for em_wdtcorner in EM_this_tb:
                                    for file in os.listdir(dirpath):
                                        if (
                                            "wdt.summary_" in file and "~" not in file
                                        ):  # reads SHE parameters
                                            if str(file[-1]) not in EM_this_tb:
                                                # print(file)
                                                with open(file, "r+") as f_wdt:

                                                    if os.stat(file).st_size != 0:
                                                        next(f_wdt)

                                                    for line in f_wdt:

                                                        new_line = line.split(" ")
                                                        splitted_nl = list(filter(None, new_line))
                                                        # print('hi',splitted_nl)

                                                        res = [
                                                            s for s in splitted_nl if element in s
                                                        ]  # check if element is a substring to any of the strings in the list
                                                        # print(res)
                                                        listToStr = " ".join(
                                                            res
                                                        )  # convert to string for if comparison below
                                                        # print('oo',listToStr)

                                                        if (
                                                            listToStr == splitted_nl[3]
                                                        ):  # print dTmos parameter
                                                            splitted_nl[4] = pd.to_numeric(
                                                                splitted_nl[4]
                                                            )
                                                            # print(file)
                                                            em_dtmos.append(splitted_nl[4])
                                                            # print(em_dtmos)
                                                            existed_element.append(element)
                                                            # print('em dtmos:',str(em_dtmos))
                                                            for c in em_dtmos:
                                                                if len(em_dtmos) == 1:
                                                                    max_emdtmos = c

                                                            if len(em_dtmos) > 1:
                                                                max_emdtmos = max(em_dtmos)

                                                            sheet.cell(
                                                                row=i + 1,
                                                                column=column_counter + 1,
                                                                value=max_emdtmos,
                                                            )
                                                            # 	print('dtmos',i,' ',column_counter)
                                                            # print('em_dtmos '+str(max_emdtmos))

                                                            existed_element.append(element)

                                                            if max_emdtmos:
                                                                for (
                                                                    header
                                                                ) in (
                                                                    measurements_headers
                                                                ):  # print dTmos instance and label in comments section
                                                                    if (
                                                                        not isinstance(
                                                                            header, float
                                                                        )
                                                                        and "Comments" in header
                                                                    ):

                                                                        comments = [
                                                                            "dTmos instance:",
                                                                            splitted_nl[5],
                                                                            "dTmos model:",
                                                                            splitted_nl[6],
                                                                        ]
                                                                        stringcomments = " ".join(
                                                                            comments
                                                                        )
                                                                        # print(stringcomments)
                                                                        sheet.cell(
                                                                            row=i + 1,
                                                                            column=measurements_headers.index(
                                                                                header
                                                                            )
                                                                            + 1,
                                                                            value=stringcomments,
                                                                        )

                                                                        break

                                                        if (
                                                            listToStr == splitted_nl[2]
                                                        ):  # print dTod parameters
                                                            splitted_nl[7] = pd.to_numeric(
                                                                splitted_nl[7]
                                                            )
                                                            em_dtod.append(splitted_nl[7])
                                                            existed_element.append(element)

                                                            for d in em_dtod:
                                                                if len(em_dtod) == 1:
                                                                    max_emdtod = d

                                                            if len(em_dtod) > 1:
                                                                max_emdtod = max(em_dtod)

                                                            sheet.cell(
                                                                row=i + 1,
                                                                column=column_counter + 1,
                                                                value=max_emdtod,
                                                            )
                                                            # print('em_dtod '+str(max_emdtod))
                                                            # print('dtod',i,' ',column_counter)

                                                            break

                                    emvalue = []
                                    irvalue = []
                                    for file in os.listdir(dirpath):
                                        file = file.replace(
                                            "_vmax.summary_", "mv" + "_vmax.summary_"
                                        )  # (old/new)

                                        if not isinstance(element, float):
                                            element = (
                                                element.lower()
                                            )  # the files names are only in lower case
                                            element = element.split()
                                            element = "".join(element)
                                            # print('element',element)

                                            if element in file:

                                                if (
                                                    "vmax" not in file
                                                    and "~" not in file
                                                    and element in file
                                                    and ".summary_" in file
                                                    and "_i" in file
                                                    or (
                                                        element in file
                                                        and "acpc" in file
                                                        and ".summary_" in file
                                                        and "~" not in file
                                                    )
                                                ):  # grab the I/Imax to fill each column, EM summary

                                                    if str(file[-1]) not in EM_this_tb:
                                                        # print(file)

                                                        EM_onlist.append(file[-1])

                                                        with open(file, "r") as f:

                                                            first_line = (
                                                                f.readline()
                                                            )  # writes the first line
                                                            new_line = first_line.split(" ")
                                                            new_line = list(filter(None, new_line))

                                                            second_line = (
                                                                f.readline()
                                                            )  # writes the second line
                                                            new_second_line = second_line.split(" ")
                                                            new_second_line = list(
                                                                filter(None, new_second_line)
                                                            )
                                                            new_second_line[3] = pd.to_numeric(
                                                                new_second_line[3]
                                                            )

                                                            emvalue.append(new_second_line[3])
                                                            existed_element.append(element)

                                                            for z in emvalue:
                                                                if len(emvalue) == 1:
                                                                    max_emvalue = z

                                                            if len(emvalue) > 1:
                                                                max_emvalue = max(emvalue)

                                                            sheet.cell(
                                                                row=i + 1,
                                                                column=column_counter + 1,
                                                                value=max_emvalue,
                                                            )
                                                            # print('em',i,' ',column_counter)
                                                            # print('EM'+str(max_emvalue))

                                                if (
                                                    element in file
                                                    and "vmax.summary_" in file
                                                    and "~" not in file
                                                ):  # grab IR drop to fill each column,located in the first lime.first column, IR summary
                                                    if str(file[-1]) not in IR_this_tb:
                                                        file = file.replace(
                                                            "mv" + "_vmax.summary_",
                                                            "_vmax.summary_",
                                                        )
                                                        layerlist.append(element)
                                                        existed_element.append(element)
                                                        # print(file)
                                                        IR_onlist.append(file[-1])
                                                        with open(file, "r") as lf:
                                                            for line in lf:
                                                                if (
                                                                    (
                                                                        layer_select in line
                                                                        or layer_select + "_"
                                                                        in line
                                                                        or layer_select + "/"
                                                                        in line
                                                                    )
                                                                    and layer_select
                                                                    in line.split()[2]
                                                                ):

                                                                    line = line.split(" ")
                                                                    line = list(filter(None, line))
                                                                    line[1] = pd.to_numeric(line[1])

                                                                    irvalue.append(line[1])
                                                                    layerlist.remove(element)

                                                                    for y in irvalue:
                                                                        if len(irvalue) == 1:
                                                                            max_irvalue = y

                                                                    if len(irvalue) > 1:
                                                                        max_irvalue = max(irvalue)

                                                                    sheet.cell(
                                                                        row=i + 1,
                                                                        column=column_counter + 1,
                                                                        value=max_irvalue,
                                                                    )
                                                                    print(
                                                                        "IR", i, " ", column_counter
                                                                    )
                                                                    # print('IR '+str(max_irvalue))
                                                                    break

                                    column_counter += 1  # goes throgh columns

                                existed_element = set(existed_element)

                                lower_measure = [x.lower() for x in measurements]
                                lower_measure = set(lower_measure)
                                unfound_element = list(set(lower_measure) - set(existed_element))

                                unfound_EM = []
                                unfound_IR = []
                                if len(unfound_element) > 0:
                                    unfound_element = ",".join(unfound_element)

                                    for paramEM in EM_param:
                                        if paramEM.lower() in unfound_element:
                                            unfound_EM.append(paramEM)

                                    for paramIR in IR_param:
                                        if paramIR.lower() in unfound_element:
                                            unfound_IR.append(paramIR)

                                    unfound_EM = ",".join(unfound_EM)
                                    unfound_IR = ",".join(unfound_IR)

                                    EM_onlist = list(set(EM_onlist))
                                    EM_onlist.sort()
                                    IR_onlist = list(set(IR_onlist))
                                    IR_onlist.sort()
                                    EM_onlist = ",".join(EM_onlist)
                                    IR_onlist = ",".join(IR_onlist)

                                    EM_this_tb = ",".join(EM_this_tb)
                                    IR_this_tb = ",".join(IR_this_tb)

                                    if len(EM_onlist) == 0:
                                        errormsgs.write(
                                            "WARNING: For testbench "
                                            + tb_name
                                            + ", there is no EM corner left to pick from after excluding EM corners: ["
                                            + EM_this_tb
                                            + "], in path: "
                                            + dirpath
                                            + "\n"
                                        )
                                    else:
                                        if len(EM_this_tb) == 0:
                                            errormsgs.write(
                                                "For testbench "
                                                + tb_name
                                                + ", EM extracted data from found corners ["
                                                + str(EM_onlist)
                                                + "], no EM corners are excluded.\nWARNING: "
                                                + unfound_EM.lower()
                                                + " is/are not available in path: "
                                                + dirpath
                                            )
                                        else:
                                            errormsgs.write(
                                                "For testbench "
                                                + tb_name
                                                + ", EM extracted data from found corners ["
                                                + str(EM_onlist)
                                                + "] and excluded corners ["
                                                + EM_this_tb
                                                + "].\nWARNING: "
                                                + unfound_EM.lower()
                                                + " is/are not available in path: "
                                                + dirpath
                                            )
                                        errormsgs.write("\n")

                                    if len(IR_onlist) == 0:
                                        errormsgs.write(
                                            "WARNING: For testbench "
                                            + tb_name
                                            + ", there is no IR corner left to pick from after excluding IR corners: ["
                                            + IR_this_tb
                                            + "], in path: "
                                            + dirpath
                                            + "\n"
                                        )
                                    else:
                                        if len(IR_this_tb) == 0:
                                            errormsgs.write(
                                                "For testbench "
                                                + tb_name
                                                + ", IR extracted data from found corners ["
                                                + str(IR_onlist)
                                                + "], no IR corners are excluded.\nWARNING: "
                                                + unfound_IR.lower()
                                                + " is/are not available in path: "
                                                + dirpath
                                            )
                                        else:
                                            errormsgs.write(
                                                "For testbench "
                                                + tb_name
                                                + ", IR extracted data from found corners ["
                                                + str(IR_onlist)
                                                + "] and excluded corners ["
                                                + IR_this_tb
                                                + "].\nWARNING: "
                                                + unfound_IR.lower()
                                                + " is/are not available in path: "
                                                + dirpath
                                            )
                                        errormsgs.write("\n")

                                if len(layerlist) > 0:
                                    layerlist = ",".join(layerlist)

                                    errormsgs.write(
                                        "WARNING: For tetbench "
                                        + tb_name
                                        + ", "
                                        + layerlist
                                        + " with "
                                        + layer_select
                                        + " layer is/are not found in path: "
                                        + dirpath
                                    )
                                    errormsgs.write("\n")

                        else:
                            errormsgs.write(
                                "WARNING: testbench: {"
                                + tb_name
                                + "} in the Excel file does not exist in local directory"
                            )
                            errormsgs.write("\n")
                        errormsgs.write("\n")

            i += 1

        os.chdir(cwd)

        f3 = open("uploadcheck.txt", "w+")
        # checking if all the given macro names were available in excelsheet
        check_list = set(check_list)
        macro_list = set(macro_list)
        not_foundlist = macro_list.symmetric_difference(
            check_list
        )  # new set with elements in either checklist or macrolist but not both
        not_found_macro = ",".join(not_foundlist)

        if len(not_foundlist) >= len(macro_list):
            errormsgs.write(
                "WARNING: macro {" + not_found_macro + "} is/are not found in: " + file_name
            )
            errormsgs.write("\n")
            print("Please check erroroutput.log for any warnings and errors")
            sys.exit()

        else:
            if len(not_foundlist) < len(macro_list) and len(not_foundlist) != 0:
                errormsgs.write(
                    "WARNING: macro {" + not_found_macro + "} is/are not found in: " + file_name
                )
                errormsgs.write("\n")
            os.system("chmod 777 " + os.getcwd() + "/" + file_name)
            wb.save(os.getcwd() + "/" + file_name)
            print("\n")

            if using_p4[0]:
                command = "p4 edit " + p4_root[1]
                os.system("cp " + file_name + " " + os.path.dirname(p4_root[1]))
                subprocess.check_output(command, shell=True)
                subprocess.check_output('p4 submit -d "Uploading"', shell=True)

            else:
                destination_path = file_path.replace(file_name, "")
                up = (
                    "curl --ntlm --user "
                    + username
                    + ":"
                    + "'"
                    + pwd
                    + "'"
                    + " --head --fail -k --upload-file "
                    + os.getcwd()
                    + "/"
                    + file_name
                    + " "
                    + destination_path
                    + ' --output "uploadcheck.txt" '
                )
                os.system(up)  # UPLOADING FILE TO THE DESTINATION PATH

                f3.close()

                with open("uploadcheck.txt", "r+") as f3:
                    checksuccess = 0
                    for line in f3.readlines():
                        if "200" in line:
                            checksuccess = checksuccess + 1
                            break
                if checksuccess == 1:
                    print("The output Excel file is uploaded here successfully:", file_path)

                else:
                    print(
                        "You do not have permission to upload the file or your file did no get downloaded successfully"
                    )
                    print(
                        "Note:If you do not have editing privileges to this SharePoint path, you cannot upload the output file"
                    )

    print("Please check erroroutput.log for any warnings and errors\n")

    # delete the generated outptut xlsx file in case the system exits before reaching here
    test = os.listdir(cwd)
    for item in [x for x in test if x.startswith(file_name)]:
        os.remove(os.path.join(cwd, item))


EM_gui_list = []
IR_gui_list = []
EM_entrylist = []
IR_entrylist = []
IRcornerlist = []
EMcornerlist = []


def emir_summary():
    msg1 = tk.Label(
        scrollable_frame,
        text="Please enter exclude corners (comma separated) below, for example:2,5 ",
        justify=tk.LEFT,
        font=14,
        padx=20,
    )
    msg1.pack(anchor="w")
    tk.Label(
        scrollable_frame,
        text='If you want to skip or ignore this testbench, please enter "all" for both EM and IR testbenches',
        justify=tk.LEFT,
        font=14,
        padx=20,
    ).pack()
    tk.Label(
        scrollable_frame,
        text="If you want to include all the corners for this testbench, please just leave the entry bar blank",
        justify=tk.LEFT,
        font=14,
        padx=20,
    ).pack()

    for i in range(len(tblist_gui)):

        macro_frame = tk.Frame(scrollable_frame)
        macro_frame.pack(anchor="w", expand=True)
        tb_frame = tk.Frame(scrollable_frame)
        tb_frame.pack(anchor="w", expand=True)
        EMIR_frame = tk.Frame(scrollable_frame)
        EMIR_frame.pack(anchor="w", expand=True)

        EM_corner_ss = tk.StringVar()
        EM_gui_list.append(EM_corner_ss)

        IR_corner_ss = tk.StringVar()

        if i == 0:
            macro_lable = tk.Label(
                macro_frame,
                text="\nMacro: " + macrolist_gui[i],
                justify=tk.LEFT,
                font=14,
                anchor="w",
                padx=20,
            )
            macro_lable.pack(side="left")
            tk.Label()
        else:
            if macrolist_gui[i] != macrolist_gui[i - 1]:
                macro_lable = tk.Label(
                    macro_frame,
                    text="\nMacro:" + macrolist_gui[i],
                    justify=tk.LEFT,
                    font=14,
                    anchor="w",
                    padx=20,
                )
                macro_lable.pack(side="left")

        tk.Label(
            tb_frame,
            text="\nTB: " + tblist_gui[i] + ", " + str(modelist_gui[i]) + "",
            justify=tk.LEFT,
            anchor="w",
            font=14,
            padx=20,
        ).pack()

        EM_label = tk.Label(
            EMIR_frame, text="EM Exclude: ", justify=tk.LEFT, anchor="w", font=14, padx=20
        )
        EM_label.pack(side="left")
        myEntry = tk.Entry(EMIR_frame, textvariable=EM_corner_ss, font=14, width=20)
        myEntry.focus_set()
        myEntry.pack(side="left", fill=tk.X)
        EM_entrylist.append(myEntry)

        IR_label = tk.Label(
            EMIR_frame, text="IR Exclude: ", justify=tk.LEFT, anchor="w", font=14, padx=20
        )
        IR_label.pack(side="left")
        myEntry2 = tk.Entry(EMIR_frame, textvariable=IR_corner_ss, font=14, width=20)
        myEntry2.focus_set()
        myEntry2.pack(side="left", fill=tk.X)
        IR_entrylist.append(myEntry2)

    def print_entry():
        for entry in EM_entrylist:
            # print(entry.get())
            EMcornerlist.append(entry.get())

        for entry in IR_entrylist:
            # print(entry.get())
            IRcornerlist.append(entry.get())

    global enterEntry_submit
    enterEntry_submit = tk.Button(
        scrollable_frame,
        text="Submit",
        font=14,
        command=combine_funcs(print_entry, checkandprintmacros),
    )
    enterEntry_submit.pack(anchor=tk.W)
    if len(sys.argv) > 1:
        root.mainloop()


def layer():
    tk.Label(
        scrollable_frame,
        text=" Metal Layer name, for example: m3      ",
        justify=tk.LEFT,
        font=14,
        padx=20,
    ).pack()
    myEntry = tk.Entry(scrollable_frame, textvariable=layer_s, font=14, width=20)
    myEntry.focus_set()

    myEntry.pack(expand=True, fill="x")
    enterEntry = tk.Button(
        scrollable_frame,
        text="Enter",
        font=14,
        command=combine_funcs(get_layer, find_tb, emir_summary),
    )
    enterEntry.pack(anchor=tk.W)


def macros():

    tk.Label(
        scrollable_frame,
        text="Macro names comma seperated:         ",
        justify=tk.LEFT,
        font=14,
        padx=20,
    ).pack()
    global myEntry3
    myEntry3 = tk.Entry(scrollable_frame, textvariable=macro_s, font=14, width=20)
    myEntry3.focus()
    myEntry3.pack(expand=True, fill="x")
    global enterEntry_macro
    enterEntry_macro = tk.Button(
        scrollable_frame, text="Enter", font=14, command=combine_funcs(get_macro, layer)
    )
    enterEntry_macro.pack(anchor=tk.W)


username = getpass.getuser()


labelcounter1 = 0
labelcounter2 = 0
labelcounter3 = 0
countpw = 0


def check_pwds():
    global countpw, labelcounter1, labelcounter2, labelcounter3
    cleaned_path = get_entry.string.split(".xlsx")[
        0
    ]  # it cutes everything after .xlsx including itself
    cleaned_path = cleaned_path + ".xlsx"
    attemptpw = password.get()
    cleaned_attemptpw = "\\" + "\\".join(attemptpw)

    if cleaned_path.startswith("//"):

        using_p4[0] = True

        print("downloading from perforce")
        command = "p4 sync -f " + cleaned_path
        locate_p4 = "p4 where " + cleaned_path
        subprocess.check_output(command, shell=True)
        p4_path = subprocess.check_output(locate_p4, shell=True, universal_newlines=True)

        xlsx_path = str(p4_path.split()[2])
        os.system("cp " + xlsx_path + " .")

        macros()

    else:
        check = (
            "curl --ntlm -k --user "
            + username
            + ":"
            + cleaned_attemptpw
            + " --head --fail "
            + cleaned_path
            + ' --output "errorcheck.txt" '
        )
        f = open("errorcheck.txt", "w")
        os.system(check)
        f.close()
        global errorlabel1, errorlabel2, errorlabel3
        with open("errorcheck.txt", "r") as f:
            a = 0
            if os.stat("errorcheck.txt").st_size == 0 and countpw == 0:
                errorlabel1 = tk.Label(
                    scrollable_frame,
                    text="Invalid URL\nPlease re-enter the password and the correct URL of the EMIR Excel file ",
                    justify=tk.LEFT,
                    font=13,
                    padx=20,
                )
                errorlabel1.pack()

                labelcounter1 = labelcounter1 + 1
                print(
                    "ERROR:Invalid URL, please re-enter the password and the correct URL of the EMIR Excel file "
                )
                delete_entry()
                enterEntry.pack_forget()
            for _ in [x for x in f.readlines() if "200 OK" in x]:
                a = a + 1
                if labelcounter1 == 1:
                    errorlabel1.destroy()
                if labelcounter2 == 1:
                    errorlabel2.destroy()
                if labelcounter3 == 1:
                    errorlabel3.destroy()

                macros()

        with open("errorcheck.txt", "r") as f2:
            i = len([x for x in f2.readlines() if "401 Unauthorized" in x])

        countpw = countpw + 1

        if i == 2 and countpw == 1:
            delete_entry()
            enterEntry.pack_forget()
            errorlabel2 = tk.Label(
                scrollable_frame,
                text="Invalid password or/and URL\nPlease re-enter the password and the correct URL of the EMIR Excel file ",
                justify=tk.LEFT,
                font=13,
                padx=20,
            )
            errorlabel2.pack()
            labelcounter2 = labelcounter2 + 1
            print(
                "ERROR:Invalid password or/and url, please re-enter the password and the correct URL of the EMIR Excel file "
            )

        if i == 1 and a == 0 and countpw == 1:
            errorlabel3 = tk.Label(
                scrollable_frame,
                text="URL does not exist\nPlease re-enter the password and the correct URL of the EMIR Excel file ",
                justify=tk.LEFT,
                font=13,
                padx=20,
            )
            print(
                "ERROR:URL does not exist, please re-enter the password and the correct URL of the EMIR Excel file "
            )
            errorlabel3.pack()
            labelcounter3 = labelcounter3 + 1
            delete_entry()
            enterEntry.pack_forget()

        if countpw == 2 and a == 0:
            tk.Label(
                scrollable_frame,
                text="WARNING:Enough password attempts, quit now to prevent lockout.\nIf you believe the password is correct, re-check the URL of the EMIR Excel file ",
                justify=tk.LEFT,
                font=13,
                padx=20,
            ).pack()
            # 	  		 print("WARNING:Enough password attempts,if you believe the password is correct, recheck the path")
            print(
                "WARNING:Too many password attempts. Closing to prevent account lockout. If you believe the password is correct, recheck the URL of the EMIR Excel file "
            )
            root.destroy()


def pwds():
    global enterEntry2
    tk.Label(
        scrollable_frame,
        text="Enter password for user:" + username + "         ",
        justify=tk.LEFT,
        font=14,
        padx=20,
    ).pack()
    global myEntry2
    myEntry2 = tk.Entry(scrollable_frame, textvariable=password, show="*", font=14, width=20)
    myEntry2.focus()
    myEntry2.pack(expand=True, fill="x")
    enterEntry2 = tk.Button(
        scrollable_frame,
        text="Enter",
        font=14,
        command=combine_funcs(get_pw, check_pwcounter, check_pwds),
    )
    enterEntry2.pack(anchor=tk.W)


def help():
    print(
        '\n- The purpose of this script is to auto-fill the template that is inputted to the script with the worst case EMIR and SHE data found from EMIR simulated corners.\n- The script will first run alphaXaSummary and alphaXaWdtSummary files to generate summary files for every ascii file found in the TB directory.\n- Then it will sweep through all the corners (except the excluded ones) and will auto-fill the EMIR template with the worst case.\n- EM and IR exclude corners are asked for each testbench.\n- Note: ensure each testbench has a unique <mode>. Because <mode> is an important distinguisher if you have two or more exact <testbench> names under one <macro> but want different exclude corners for each testbench.\n- If you want to skip or ignore a testbench, plese enter "all" for both EM and IR exclude corners.\n- If you want to include all the corners for this testbench, please just leave the entry blank.\n'
    )

    print(
        "Usage:\n 	/remote/cad-rep/projects/alpha/alpha_common/bin/EmirDataExtractin.py   -config <config file.ini> (optional) [--help|-h]"
    )
    print("\nDescription:")
    print(
        "	-config <file.ini> = Create a config file for user input (please see below for sample config file) "
    )
    print("	--help|-h= will print script usage info")
    print("\nExamples:")
    print(
        "	EmirDataExtractin.py <--- GUI mode: Running without any options will open a GUI for the user to provide inputs to the script"
    )
    print("	EmirDataExtractin.py -config <config file.ini>")
    print("	EmirDataExtractin.py --help \n\n")

    print("***************************************")
    print("** SAMPLE CONFIG FILE  **")
    print("***************************************")
    print("In the user config, 3 parameters must defined as follows:\n")
    print(
        "	1. Sharepoint path, this is the path to the template to be auto-filled, this file has to exist prior to uploading "
    )
    print("	2. Macro name, this is the macros that you wish to extract the EMIR and the SHE from")
    print(
        "	3. Layer name, this is the metal layer, for example, m3 that you wish to extract the IR drop from"
    )

    print("Sample config file for a new template:")
    print(
        "	sharepoint_path = https://sp-sg/sites/msip-eddr/proj/eddra/ckt/Shared%20Documents/CAD%20Docs/EMIR-SHE/emirtest.xlsx"
    )
    print("	macro_name=dwc_ddrphy_rxac,dwc_ddrphy_rxdq,dwc_ddrphy_lcdl")
    print("	metal_layer=m1\n")

    print(
        "If you run the config option, EM and IR exclude corners will be asked later for each testbench in the command line.\n"
    )


if len(sys.argv) == 1:
    tk.Label(
        scrollable_frame, text="Path to Excel sheet:       ", justify=tk.LEFT, font=14, padx=20
    ).pack(expand=True, fill="x", side="top")
    myEntry = tk.Entry(scrollable_frame, textvariable=path, font=14, width=35)
    myEntry.focus_set()
    myEntry.pack(expand=True, fill="x")

    enterEntry = tk.Button(
        scrollable_frame, text="Enter", font=14, command=combine_funcs(get_entry, check_pathcounter)
    )
    enterEntry.pack(anchor=tk.W)

else:
    if len(sys.argv) > 1 and sys.argv[1] == "-config":
        checkandprintmacros()

        # checkandprintmacros(my_name,my_macro_list,my_layer_select)
        # if len(sys.argv)>1:
        # 	checkandprintmacros(my_name,my_macro_list,my_layer_select)

    if sys.argv[1] == "--help" or sys.argv[1] == "-h":
        help()
        sys.exit()

if len(sys.argv) == 1:
    root.mainloop()
